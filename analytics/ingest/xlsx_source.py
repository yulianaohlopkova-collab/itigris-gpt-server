from __future__ import annotations

import time
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import io
import zipfile
import re

from ..contracts import KPIBlock, MonthFacts, MonthMeta, WeekMeta


def _norm(s: Any) -> str:
    t = str(s or "").strip()
    t = t.replace("\u00a0", " ")
    while "  " in t:
        t = t.replace("  ", " ")
    return t


def _norm_l(s: Any) -> str:
    return _norm(s).lower()


def _cell(ws: Any, r: int, c: int) -> str:
    return _norm(ws.cell(r, c).value)


def _row_values(ws: Any, r: int, min_c: int, max_c: int) -> List[str]:
    return [_cell(ws, r, c) for c in range(min_c, max_c + 1)]


def _is_blank_row(vals: List[str]) -> bool:
    return not any(v for v in vals)


def _looks_like_kpi_title(s: str) -> bool:
    t = _norm(s)
    if not t or len(t) < 4:
        return False
    # Most KPI blocks are uppercase in XLS exports.
    if any(ch.isalpha() for ch in t) and (t.upper() == t):
        return True
    # Some sheets use mixed-case section titles (e.g. "ОПРАВЫ СТМ, 7 салонов").
    tl = t.lower()
    if tl.startswith(("оправы", "линзы", "солнцезащит", "контактные линзы", "аксессуары")):
        if any(k in tl for k in [" стм", "продан", "фотохром", "по бренду", "по дизайну", "по типу", "по целевой"]):
            return True
    return False


def extract_month_meta(ws: Any, month_label: str) -> MonthMeta:
    anchor_r: Optional[int] = None
    for r in range(1, 80):
        if "технические данные" in _norm_l(_cell(ws, r, 1)):
            anchor_r = r
            break
    if anchor_r is None:
        raise ValueError("month_meta_anchor_not_found")

    labels = _row_values(ws, anchor_r, 1, 80)
    values = _row_values(ws, anchor_r + 1, 1, 80)

    def idx_of(substr: str) -> Optional[int]:
        for i, lab in enumerate(labels):
            if substr in _norm_l(lab):
                return i
        return None

    i_start = idx_of("начало месяца")
    i_end = idx_of("конец месяца")
    i_days = idx_of("кол-во дней")
    if i_start is None or i_end is None:
        raise ValueError("month_meta_missing_start_end")

    start = values[i_start]
    end = values[i_end]
    days = int(float(values[i_days])) if i_days is not None and values[i_days] else 0

    weeks: List[WeekMeta] = []
    for w in ["i", "ii", "iii", "iv", "v"]:
        i_ws = idx_of(f"{w} неделя начало")
        i_we = idx_of(f"{w} неделя конец")
        i_wd = idx_of(f"{w} неделя дней")
        if i_ws is None or i_we is None:
            continue
        w_start = values[i_ws]
        w_end = values[i_we]
        w_days = int(float(values[i_wd])) if i_wd is not None and values[i_wd] else 0
        weeks.append(WeekMeta(name=w.upper(), start=w_start, end=w_end, days=w_days))

    return MonthMeta(month=month_label, start=start, end=end, days=days, weeks=weeks)


def extract_kpi_blocks(ws: Any, header_marker: str = "салон / показатель", advance_on_block: bool = True) -> List[KPIBlock]:
    max_r = 800
    max_c = 80
    blocks: List[KPIBlock] = []
    r = 1
    while r <= max_r:
        first = _cell(ws, r, 1)
        if _looks_like_kpi_title(first):
            title = first
            header_r: Optional[int] = None
            for rr in range(r + 1, min(r + 12, max_r) + 1):
                if header_marker in _norm_l(_cell(ws, rr, 1)):
                    header_r = rr
                    break
            if header_r is None:
                r += 1
                continue

            header = _row_values(ws, header_r, 1, max_c)
            while header and not header[-1]:
                header.pop()

            data: List[List[str]] = []
            rr = header_r + 1
            while rr <= max_r:
                vals = _row_values(ws, rr, 1, len(header))
                if _is_blank_row(vals):
                    if data:
                        break
                    rr += 1
                    continue
                if header_marker in _norm_l(vals[0]):
                    break
                # Some tables have uppercase values in the first column (e.g., brand names).
                if _looks_like_kpi_title(vals[0]) and data and _is_blank_row(vals[1:]):
                    break
                data.append(vals)
                rr += 1

            if data:
                blocks.append(KPIBlock(title=title, header=header, rows=data))
                if advance_on_block:
                    r = rr
                    continue
        r += 1
    return blocks


def pick_kpis(blocks: List[KPIBlock]) -> Dict[str, KPIBlock]:
    def score(title: str, needles: List[str]) -> int:
        tl = title.lower()
        return sum(1 for n in needles if n in tl)

    wanted = {
        "revenue": ["выруч", "показатели продаж"],
        "avg_order_check": ["средний чек", "ср чек", "средняя сумма", "чек заказа", "чек на очки"],
        "avg_income_per_client": ["средний доход", "доход от клиента"],
        "conversion": ["конвер"],
    }

    out: Dict[str, KPIBlock] = {}
    for key, needles in wanted.items():
        best: Tuple[int, Optional[KPIBlock]] = (0, None)
        for b in blocks:
            sc = score(b.title, needles)
            if sc > best[0]:
                best = (sc, b)
        if best[1] is not None and best[0] > 0:
            out[key] = best[1]
    return out


def _pick_first(blocks: List[KPIBlock], needles: List[str]) -> Optional[KPIBlock]:
    def score(title: str) -> int:
        tl = title.lower()
        return sum(1 for n in needles if n in tl)

    best_sc = 0
    best: Optional[KPIBlock] = None
    for b in blocks:
        sc = score(b.title)
        if sc > best_sc:
            best_sc = sc
            best = b
    return best if best_sc > 0 else None


def pick_frames_mix_blocks(blocks: List[KPIBlock]) -> Dict[str, KPIBlock]:
    out: Dict[str, KPIBlock] = {}
    stm = _pick_first(blocks, ["оправы стм"])
    if stm:
        out["stm_units"] = stm
    by_brand = _pick_first(blocks, ["проданные оправы по бренду"])
    if by_brand:
        out["by_brand"] = by_brand
    by_design = _pick_first(blocks, ["проданные оправы по дизайну"])
    if by_design:
        out["by_design"] = by_design
    by_type = _pick_first(blocks, ["проданные оправы по типу"])
    if by_type:
        out["by_type"] = by_type
    by_target = _pick_first(blocks, ["проданные оправы по целевой"])
    if by_target:
        out["by_target_group"] = by_target
    avg_check = _pick_first(blocks, ["средний чек", "ср. чек", "ср чек", "чек"])
    if avg_check:
        out["avg_check"] = avg_check
    return out


def pick_lenses_mix_blocks(blocks: List[KPIBlock]) -> Dict[str, KPIBlock]:
    out: Dict[str, KPIBlock] = {}
    # Lenses sheet often uses "ОЛ" wording.
    by_manuf = _pick_first(blocks, ["по созданным", "производ"])
    if by_manuf:
        out["by_manufacturer"] = by_manuf
    by_brand = _pick_first(blocks, ["по бренду"])
    if by_brand:
        out["by_brand"] = by_brand
    photo = _pick_first(blocks, ["фотохром"])
    if photo:
        out["photochromic"] = photo
    return out


def _parse_lenses_pdf_mix(pdf_path: str) -> Dict[str, KPIBlock]:
    """
    PDF parsing is intentionally disabled in analytics v1.
    It's brittle and we prefer XLS values, with a ZIP/HTML fallback when needed.
    """
    _ = pdf_path
    return {}


def _parse_lenses_zip_html_mix(zip_path: str, month_label: str) -> Dict[str, Any]:
    """
    Preferred v1 fallback for lenses: parse Google-Sheets HTML export from a zip.
    Returns a dict with:
      - source: {type,id,entry}
      - blocks: {by_manufacturer_brand, by_department_brand, photochromic_by_department_brand}
    """
    entry = f"линзы {month_label}.html"
    try:
        with zipfile.ZipFile(zip_path, "r") as z:
            if entry not in z.namelist():
                return {}
            html = z.read(entry).decode("utf-8", "ignore")
    except Exception:
        return {}

    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if table is None:
        return {}

    def parse_table_grid(tbl) -> List[List[str]]:
        grid: List[List[str]] = []
        spans: Dict[Tuple[int, int], Tuple[int, str]] = {}

        rows = tbl.find_all("tr")
        for r_i, tr in enumerate(rows):
            cells = tr.find_all(["td", "th"])
            row: List[str] = []
            c_i = 0

            def take_span() -> Optional[str]:
                nonlocal c_i
                v = spans.get((r_i, c_i))
                if not v:
                    return None
                remaining, val = v
                if remaining > 1:
                    spans[(r_i + 1, c_i)] = (remaining - 1, val)
                del spans[(r_i, c_i)]
                return val

            for cell in cells:
                while True:
                    sv = take_span()
                    if sv is None:
                        break
                    row.append(sv)
                    c_i += 1

                txt = cell.get_text(" ", strip=True).replace("\u00a0", " ")
                rs = int(cell.get("rowspan") or 1)
                cs = int(cell.get("colspan") or 1)
                for _ in range(cs):
                    row.append(txt)
                if rs > 1:
                    for k in range(cs):
                        spans[(r_i + 1, c_i + k)] = (rs - 1, txt)
                c_i += cs

            while True:
                sv = take_span()
                if sv is None:
                    break
                row.append(sv)
                c_i += 1

            while row and not row[-1]:
                row.pop()
            grid.append([_norm(x) for x in row])
        return grid

    grid = parse_table_grid(table)
    if len(grid) < 6:
        return {}

    # Find a header row near the top. In 05.2026 exports it is usually row 3 (0-based),
    # but we search for a known token to be more robust.
    r_header = None
    for i in range(min(80, len(grid))):
        if grid[i] and "салон / показатель" in _norm_l(grid[i][0]):
            r_header = i
            break
    if r_header is None:
        r_header = 3

    def cell(r: int, c: int) -> str:
        if r < 0 or r >= len(grid):
            return ""
        row = grid[r]
        if c < 0 or c >= len(row):
            return ""
        return row[c]

    manuf_cols = list(range(1, 7))   # B..G
    dept_cols = list(range(9, 15))   # J..O
    photo_cols = list(range(17, 23)) # R..W

    def block_from_cols(title: str, first_col_name: str, cols: List[int]) -> KPIBlock:
        header = [first_col_name, "brand", "sum_rub", "share_pct", "count", "avg_rub"]
        rows: List[List[str]] = []
        last_first = ""
        for r in range(r_header + 1, len(grid)):
            vals = [cell(r, c) for c in cols]
            if not any(v for v in vals):
                continue
            first_raw = vals[0]
            first = first_raw or last_first
            if first_raw:
                last_first = first_raw
            if first_raw and _looks_like_kpi_title(first_raw) and rows:
                break
            if not first:
                continue
            fl = first.lower()
            # Skip subtotal/header-like rows from the HTML export.
            if fl.startswith(("общий", "итого", "всего")):
                continue
            rows.append([first, vals[1], vals[2], vals[3], vals[4], vals[5]])
        return KPIBlock(title=title, header=header, rows=rows)

    blocks: Dict[str, KPIBlock] = {}
    blocks["by_manufacturer_brand"] = block_from_cols(
        "ПРОДАННЫЕ ОЛ (HTML) — производители/бренды",
        "manufacturer",
        manuf_cols,
    )
    blocks["by_department_brand"] = block_from_cols(
        "ПРОДАННЫЕ ОЛ (HTML) — салоны/бренды",
        "department",
        dept_cols,
    )
    blocks["photochromic_by_department_brand"] = block_from_cols(
        "ПРОДАННЫЕ ФОТОХРОМНЫЕ ОЛ (HTML) — салоны/бренды",
        "department",
        photo_cols,
    )

    return {
        "source": {"type": "zip_html", "id": zip_path, "entry": entry},
        "blocks": blocks,
    }


def pick_people_blocks(blocks: List[KPIBlock]) -> Dict[str, KPIBlock]:
    out: Dict[str, KPIBlock] = {}

    def has_plan_fact_dev(b: KPIBlock) -> bool:
        hl = " ".join(_norm_l(x) for x in (b.header or []))
        return ("план" in hl) and ("факт" in hl) and ("отклон" in hl)

    def pick_semantic(title_needles: List[str]) -> Optional[KPIBlock]:
        candidates = [b for b in blocks if all(n in _norm_l(b.title) for n in title_needles)]
        ranked = sorted(candidates, key=lambda b: (has_plan_fact_dev(b), len(b.rows)), reverse=True)
        for b in ranked:
            if has_plan_fact_dev(b):
                return b
        return ranked[0] if ranked else None

    lenses = pick_semantic(["продан", "линз"])
    if lenses:
        out["sold_lenses_consultants"] = lenses
    frames = pick_semantic(["продан", "оправ"])
    if frames:
        out["sold_frames_consultants"] = frames
    sunglasses = pick_semantic(["солнцезащ"])
    if sunglasses:
        out["sold_sunglasses_consultants"] = sunglasses
    photo = pick_semantic(["фотохром"])
    if photo:
        out["photochromic_consultants"] = photo
    return out


@dataclass
class XlsxDashboardSource:
    """v1 ingestion source: manual XLSX export from ODL dashboard."""

    xlsx_path: str

    def load_month(self, month_label: str) -> MonthFacts:
        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True, read_only=True)
        kpi_sheet = f"показатели {month_label}"
        if kpi_sheet not in wb.sheetnames:
            raise ValueError(f"missing_sheet:{kpi_sheet}")

        ws = wb[kpi_sheet]
        meta = extract_month_meta(ws, month_label=month_label)
        blocks = extract_kpi_blocks(ws)
        picked = pick_kpis(blocks)

        mix_blocks: Dict[str, Any] = {}

        frames_sheet = f"оправы {month_label}"
        if frames_sheet in wb.sheetnames:
            ws_f = wb[frames_sheet]
            f_blocks = extract_kpi_blocks(ws_f, advance_on_block=False)
            mix_blocks["frames"] = {"sheet": frames_sheet, "blocks": pick_frames_mix_blocks(f_blocks)}

        lenses_sheet = f"линзы {month_label}"
        if lenses_sheet in wb.sheetnames:
            ws_l = wb[lenses_sheet]
            l_blocks = extract_kpi_blocks(ws_l, advance_on_block=False)
            lens_mix = pick_lenses_mix_blocks(l_blocks)
            # If XLS has no cached values, fall back to HTML export from the dashboard zip (preferred),
            # then PDF as last resort (currently unused in v1 output).
            lenses_pack: Dict[str, Any] = {"sheet": lenses_sheet, "blocks": lens_mix}
            if not lens_mix:
                zip_path = "/Users/roomofpromise/Downloads/ODL. Дашборд.zip"
                html_pack = _parse_lenses_zip_html_mix(zip_path, month_label=month_label)
                if html_pack:
                    lenses_pack["blocks"] = html_pack["blocks"]
                    lenses_pack["source"] = html_pack["source"]
            mix_blocks["lenses"] = lenses_pack

        people_blocks: Dict[str, Any] = {}
        cons_sheet = f"консопто {month_label}"
        if cons_sheet in wb.sheetnames:
            ws_c = wb[cons_sheet]
            c_blocks = extract_kpi_blocks(ws_c, header_marker="конс / период")
            people_blocks["consultants"] = {"sheet": cons_sheet, "blocks": pick_people_blocks(c_blocks)}

        doctor_sheet = f"врач {month_label}"
        if doctor_sheet in wb.sheetnames:
            ws_d = wb[doctor_sheet]
            d_blocks = extract_kpi_blocks(ws_d, advance_on_block=False)
            people_blocks["doctor"] = {"sheet": doctor_sheet, "blocks_all": [b.title for b in d_blocks]}

        now = int(time.time())
        return MonthFacts(
            month_meta=meta,
            kpi_blocks_all=blocks,
            kpi_blocks_picked=picked,
            mix_blocks=mix_blocks,
            orders_blocks={},
            people_blocks=people_blocks,
            source_type="xlsx_manual",
            source_id=self.xlsx_path,
            generated_at_unix=now,
        )
