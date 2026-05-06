from __future__ import annotations

import csv
import base64
import io
import os
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urljoin, urlparse, parse_qs
import asyncio

import httpx
import openpyxl
import xlsxwriter
import xlrd
from bs4 import BeautifulSoup
from fastapi import FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.openapi.docs import get_swagger_ui_html
from fastapi.openapi.utils import get_openapi
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field

from analytics import analyze_dataset, load_input_folder


app = FastAPI(
    title="ODL Sales Analyst MVP API",
    version="3.0.0",
    docs_url=None,
    redoc_url=None,
    openapi_url=None,
)

# Deployment/build markers (useful to confirm which commit Render is running).
BUILD_COMMIT = (
    os.getenv("RENDER_GIT_COMMIT")
    or os.getenv("GIT_COMMIT")
    or os.getenv("COMMIT_SHA")
    or ""
).strip()
AUTO_REFRESH_IMPL_VERSION = "web_pipeline_v7_uuid_pair_fallback"


APP_NAME = os.getenv("ITIGRIS_APP_NAME", "odl").strip() or "odl"
REMOTE_API_KEY = (
    os.getenv("ITIGRIS_REMOTE_API_KEY")
    or os.getenv("ITIGRIS_API_KEY")
    or ""
).strip()
EXTERNAL_API_KEY = os.getenv("ITIGRIS_EXTERNAL_API_KEY", "").strip()
ODL_SERVER_TOKEN = os.getenv("ODL_SERVER_TOKEN", "").strip()
SERVER_URL = os.getenv("PUBLIC_SERVER_URL", "https://itigris-gpt-server.onrender.com").strip()
TIMEOUT = float(os.getenv("ITIGRIS_TIMEOUT", "40"))
REMOTE_REMAINS_URL = f"https://optima.itigris.ru/{APP_NAME}/remoteRemains/list"

REMAINGOODS_AUTO_FETCH_URL_TEMPLATE = os.getenv("ITIGRIS_REMAINGOODSREPORT_URL_TEMPLATE", "").strip()
REMAINGOODS_AUTO_REFRESH_MIN_SECONDS = int(os.getenv("REMAINGOODS_AUTO_REFRESH_MIN_SECONDS", "600"))

REMAINGOODS_WEB_COOKIE = os.getenv("ITIGRIS_REMAINGOODSREPORT_WEB_COOKIE", "").strip()
REMAINGOODS_WEB_COMPANY_UUID = os.getenv("ITIGRIS_REMAINGOODSREPORT_WEB_COMPANY_UUID", "").strip()
REMAINGOODS_WEB_USER_ID = os.getenv("ITIGRIS_REMAINGOODSREPORT_WEB_USER_ID", "").strip()
REMAINGOODS_WEB_PAGE_UUID = os.getenv("ITIGRIS_REMAINGOODSREPORT_WEB_PAGE_UUID", "").strip()
REMAINGOODS_WEB_UUID_VALUE = os.getenv("ITIGRIS_REMAINGOODSREPORT_WEB_UUID_VALUE", "").strip()
REMAINGOODS_WEB_REPORT_TYPE = os.getenv("ITIGRIS_REMAINGOODSREPORT_WEB_REPORT_TYPE", "Контактные линзы").strip()
REMAINGOODS_WEB_PRICE_TYPE = os.getenv("ITIGRIS_REMAINGOODSREPORT_WEB_PRICE_TYPE", "Розничная").strip()
REMAINGOODS_WEB_DEPARTMENT_IDS = os.getenv("ITIGRIS_REMAINGOODSREPORT_WEB_DEPARTMENT_IDS", "").strip()

# remainGoodsReport module context overrides (separate from login page context).
ITIGRIS_REMAINGOODSREPORT_PAGE_UUID = os.getenv("ITIGRIS_REMAINGOODSREPORT_PAGE_UUID", "").strip()
ITIGRIS_REMAINGOODSREPORT_UUID_VALUE = os.getenv("ITIGRIS_REMAINGOODSREPORT_UUID_VALUE", "").strip()

# Web login (preferred over static cookies; cookies expire quickly).
ITIGRIS_WEB_LOGIN = os.getenv("ITIGRIS_WEB_LOGIN", "").strip()
ITIGRIS_WEB_PASSWORD = os.getenv("ITIGRIS_WEB_PASSWORD", "").strip()
ITIGRIS_WEB_KEY = os.getenv("ITIGRIS_WEB_KEY", "").strip()
ITIGRIS_WEB_VERSION_DESC = os.getenv("ITIGRIS_WEB_VERSION_DESC", "").strip()
ITIGRIS_WEB_BROWSER_DESC = os.getenv("ITIGRIS_WEB_BROWSER_DESC", "").strip()
ITIGRIS_WEB_USER_AGENT = os.getenv("ITIGRIS_WEB_USER_AGENT", "").strip()
ITIGRIS_WEB_PAGE_UUID = os.getenv("ITIGRIS_WEB_PAGE_UUID", "").strip()
ITIGRIS_WEB_UUID_VALUE = os.getenv("ITIGRIS_WEB_UUID_VALUE", "").strip()
ITIGRIS_WEB_USER_ID = os.getenv("ITIGRIS_WEB_USER_ID", "").strip()

ITIGRIS_WEB_LOGIN_URL = os.getenv("ITIGRIS_WEB_LOGIN_URL", f"https://optima.itigris.ru/{APP_NAME}/login/login").strip()


DEPARTMENTS: Dict[str, int] = {
    "Ленина": 1000000021,
    "Склад. Мобильный салон": 1000000020,
    "Мобильный салон": 1000000019,
    "Интернет-магазин Якутск": 1000000018,
    "Склад. Интернет-магазин Якутск": 1000000017,
    "Айсберг": 1000000016,
    "Качели": 1000000012,
    "Улуру": 1000000011,
    "Лермонтова": 1000000009,
    "Пояркова": 1000000008,
    "Склад ИП": 1000000007,
    "Цех": 1000000006,
    "Склад ООО": 1000000005,
    "Экспо": 1000000004,
    "Офис": 1000000003,
}

DEPARTMENT_ALIASES: Dict[str, int] = {
    "качели": 1000000012,
    "в качелях": 1000000012,
    "качелях": 1000000012,
    "тц качели": 1000000012,
    "айсберг": 1000000016,
    "в айсберге": 1000000016,
    "айсберге": 1000000016,
    "тц айсберг": 1000000016,
    "ленина": 1000000021,
    "ленина 7": 1000000021,
    "ленина, 7": 1000000021,
    "на ленина": 1000000021,
    "пояркова": 1000000008,
    "пояркова 5": 1000000008,
    "пояркова, 5": 1000000008,
    "на пояркова": 1000000008,
    "лермонтова": 1000000009,
    "лермонтова 49": 1000000009,
    "лермонтова, 49": 1000000009,
    "на лермонтова": 1000000009,
    "улуру": 1000000011,
    "улуруу": 1000000011,
    "улуруу молл": 1000000011,
    "в улуру": 1000000011,
    "экспо": 1000000004,
    "сахаэкспоцентр": 1000000004,
    "саха экспо": 1000000004,
}

GROUPS_RU: Dict[str, List[int]] = {
    "салоны": [
        1000000021,
        1000000019,
        1000000018,
        1000000016,
        1000000012,
        1000000011,
        1000000004,
        1000000009,
        1000000008,
    ],
    "склады_ип": [1000000007],
    "склады_ооо": [1000000005],
    "цех": [1000000006],
}

GROUP_ALIASES: Dict[str, str] = {
    "salons": "салоны",
    "салоны": "салоны",
    "салон": "салоны",
    "магазины": "салоны",
    "warehouse_ip": "склады_ип",
    "склад ип": "склады_ип",
    "ип": "склады_ип",
    "warehouse_ooo": "склады_ооо",
    "склад ооо": "склады_ооо",
    "ооо": "склады_ооо",
    "workshop": "цех",
    "цех": "цех",
}

CATEGORY_FILTERS: Dict[str, List[str]] = {
    "accessories": ["manufacturer", "brand", "model", "color", "material", "type"],
    "contactlenses": [
        "manufacturer",
        "name",
        "color",
        "radius",
        "diameter",
        "dioptre",
        "cylinder",
        "axis",
        "add",
        "wearingPeriod",
        "inPack",
    ],
    "glasses": ["manufacturer", "brand", "model", "color", "purpose", "material", "type", "size", "design"],
    "lenses": [
        "manufacturer",
        "brand",
        "index",
        "cover",
        "color",
        "diameter",
        "material",
        "geometry",
        "lensType",
        "lensClass",
        "technology",
        "dioptre",
        "cylinder",
        "add",
    ],
    "sunglasses": ["manufacturer", "brand", "model", "color", "purpose", "material", "type", "lensType", "design"],
}

CATEGORY_ALIASES: Dict[str, str] = {
    "accessories": "accessories",
    "аксессуар": "accessories",
    "аксессуары": "accessories",
    "contactlenses": "contactlenses",
    "contact lenses": "contactlenses",
    "контактные линзы": "contactlenses",
    "контактные": "contactlenses",
    "контакты": "contactlenses",
    "кл": "contactlenses",
    "мкл": "contactlenses",
    "мягкие контактные линзы": "contactlenses",
    "glasses": "glasses",
    "frames": "glasses",
    "frame": "glasses",
    "оправы": "glasses",
    "оправа": "glasses",
    "lenses": "lenses",
    "очковые линзы": "lenses",
    "ол": "lenses",
    "линзы": "lenses",
    "стекла": "lenses",
    "стекло": "lenses",
    "sunglasses": "sunglasses",
    "солнцезащитные": "sunglasses",
    "солнцезащитные очки": "sunglasses",
    "солнечные очки": "sunglasses",
    "солнце": "sunglasses",
    "сз": "sunglasses",
}

FILTER_FIELD_ALIASES: Dict[str, str] = {
    "add": "add",
    "аддидация": "add",
    "axis": "axis",
    "ось": "axis",
    "brand": "brand",
    "бренд": "brand",
    "color": "color",
    "цвет": "color",
    "cover": "cover",
    "покрытие": "cover",
    "cylinder": "cylinder",
    "цилиндр": "cylinder",
    "design": "design",
    "дизайн": "design",
    "форма": "design",
    "diameter": "diameter",
    "диаметр": "diameter",
    "dioptre": "dioptre",
    "диоптрии": "dioptre",
    "диоптрия": "dioptre",
    "сфера": "dioptre",
    "geometry": "geometry",
    "геометрия": "geometry",
    "inpack": "inPack",
    "inPack": "inPack",
    "в упаковке": "inPack",
    "штук в упаковке": "inPack",
    "index": "index",
    "индекс": "index",
    "lensclass": "lensClass",
    "lensClass": "lensClass",
    "класс линзы": "lensClass",
    "lenstype": "lensType",
    "lensType": "lensType",
    "тип линзы": "lensType",
    "manufacturer": "manufacturer",
    "производитель": "manufacturer",
    "material": "material",
    "материал": "material",
    "model": "model",
    "модель": "model",
    "name": "name",
    "название": "name",
    "наименование": "name",
    "purpose": "purpose",
    "target": "purpose",
    "целевая группа": "purpose",
    "целевая_группа": "purpose",
    "пол": "purpose",
    "radius": "radius",
    "radiusofcurvature": "radius",
    "radiusOfCurvature": "radius",
    "радиус": "radius",
    "радиус кривизны": "radius",
    "size": "size",
    "размер": "size",
    "technology": "technology",
    "технология": "technology",
    "type": "type",
    "тип": "type",
    "wearingperiod": "wearingPeriod",
    "wearingPeriod": "wearingPeriod",
    "срок ношения": "wearingPeriod",
    "период ношения": "wearingPeriod",
}

BREAKDOWN_FIELDS: Dict[str, List[str]] = {
    "accessories": ["type", "brand", "manufacturer"],
    "contactlenses": ["manufacturer", "name", "dioptre", "radius", "diameter", "wearingPeriod", "inPack"],
    "glasses": ["brand", "manufacturer", "material", "type", "purpose", "design"],
    "lenses": ["manufacturer", "brand", "lensType", "lensClass", "cover", "index", "dioptre"],
    "sunglasses": ["brand", "manufacturer", "material", "type", "purpose", "design"],
}


class RemainsFilteredRequest(BaseModel):
    category: str
    department_id: Optional[int] = None
    department_name: Optional[str] = None
    group: Optional[str] = None
    filters: Optional[Dict[str, Any]] = None
    min_price: Optional[float] = None
    max_price: Optional[float] = None
    price: Optional[float] = None
    return_items: bool = False
    items_limit: int = Field(default=0, ge=0, le=1000)


class SalesAnalyzeRequest(BaseModel):
    data: Optional[Dict[str, List[Dict[str, Any]]]] = None


class SalesAnalyzeCsvRequest(BaseModel):
    sales_period_csv: str = Field(..., description="CSV content for sales_period.csv (UTF-8).")
    plan_fact_csv: Optional[str] = Field(default=None, description="CSV content for plan_fact.csv (optional).")
    employees_csv: Optional[str] = Field(default=None, description="CSV content for employees.csv (optional).")
    categories_csv: Optional[str] = Field(default=None, description="CSV content for categories.csv (optional).")
    training_csv: Optional[str] = Field(default=None, description="CSV content for training.csv (optional).")


class RemainGoodsReportCsvRequest(BaseModel):
    report_csv: str = Field(..., description="CSV content of remainGoodsReport (UTF-8).")


class RemainGoodsReportBase64Request(BaseModel):
    filename: str = Field(..., description="Original filename (e.g. remainGoodsReport.xls).")
    file_base64: str = Field(..., description="Base64-encoded file bytes for .xls/.xlsx/.csv.")


class RemainGoodsReportSnapshotSetRequest(BaseModel):
    department_name: str = Field(..., description="Department/salon name, e.g. 'Ленина' or 'Ленина, 7'.")
    filename: str = Field(..., description="Original filename (e.g. remainGoodsReport (22).xls).")
    file_base64: str = Field(..., description="Base64-encoded remainGoodsReport file bytes (.xls/.xlsx/.csv).")


class RemainGoodsReportGlobalSnapshotSetRequest(BaseModel):
    filename: str = Field(..., description="Original filename (e.g. remainGoodsReport_all.xls).")
    file_base64: str = Field(..., description="Base64-encoded remainGoodsReport file bytes (.xls/.xlsx/.csv).")


class RemainGoodsReportSnapshotInfo(BaseModel):
    department_id: int
    department_name: str
    stored_at_unix: int
    expires_at_unix: int
    filename: str
    rows_count: int


REMAINGOODS_SNAPSHOT_TTL_SECONDS = int(os.getenv("REMAINGOODS_SNAPSHOT_TTL_SECONDS", str(24 * 60 * 60)))
# In-memory snapshot cache (Render dynos may restart; this is MVP-grade).
_contactlenses_report_snapshots: Dict[int, Dict[str, Any]] = {}
_contactlenses_report_global_snapshot: Optional[Dict[str, Any]] = None
_contactlenses_auto_fetch_state: Dict[str, Any] = {
    "last_attempt_unix": None,
    "last_success_unix": None,
    "last_error": None,
    "last_error_at_unix": None,
    "last_filename": None,
}


def csv_bytes_to_rows(raw: bytes) -> List[Dict[str, Any]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def csv_text_to_rows(text: str) -> List[Dict[str, Any]]:
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


async def load_csv_upload(file: Optional[UploadFile], required: bool) -> List[Dict[str, Any]]:
    if file is None:
        if required:
            raise HTTPException(status_code=400, detail="missing_required_csv")
        return []
    data = await file.read()
    if len(data) > 10_000_000:
        raise HTTPException(status_code=413, detail="csv_too_large")
    return csv_bytes_to_rows(data)


def require_auth_token(request: Request) -> Optional[JSONResponse]:
    if not ODL_SERVER_TOKEN:
        return JSONResponse(
            {"error": "server_token_not_configured", "detail": "Set ODL_SERVER_TOKEN in Render Environment."},
            status_code=500,
        )

    header_token = request.headers.get("x-odl-token") or request.headers.get("X-ODL-Token")
    query_token = request.query_params.get("token")
    if header_token == ODL_SERVER_TOKEN or query_token == ODL_SERVER_TOKEN:
        return None
    return JSONResponse({"error": "forbidden"}, status_code=403)


def normalize_category(category: str) -> str:
    return CATEGORY_ALIASES.get(category.strip().lower(), category.strip().lower())


def normalize_group(group: Optional[str]) -> Optional[str]:
    if not group:
        return None
    return GROUP_ALIASES.get(group.strip().lower(), group.strip().lower())


def normalize_department(department_id: Optional[int] = None, department_name: Optional[str] = None) -> Optional[int]:
    if department_id:
        return department_id
    if not department_name:
        return None

    raw = department_name.strip()
    if raw.isdigit():
        return int(raw)

    if raw in DEPARTMENTS:
        return DEPARTMENTS[raw]

    low = raw.lower()
    if low in DEPARTMENT_ALIASES:
        return DEPARTMENT_ALIASES[low]

    for name, dep_id in DEPARTMENTS.items():
        if name.lower() == low:
            return dep_id
    return None


def resolve_dep_ids(department_id: Optional[int], department_name: Optional[str], group: Optional[str]) -> List[int]:
    dep_id = normalize_department(department_id, department_name)
    if dep_id:
        return [dep_id]
    if department_name and not dep_id:
        raise ValueError("unknown_department")
    if group:
        normalized_group = normalize_group(group)
        if normalized_group not in GROUPS_RU:
            raise ValueError("unknown_group")
        return GROUPS_RU[normalized_group]
    return list(DEPARTMENTS.values())


def normalize_filter_field_name(field: str) -> str:
    key = field.strip()
    return FILTER_FIELD_ALIASES.get(key, FILTER_FIELD_ALIASES.get(key.lower(), key))


def normalize_filters(category: str, filters: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not filters:
        return None

    allowed = set(CATEGORY_FILTERS[category])
    normalized: Dict[str, Any] = {}
    ignored: Dict[str, Any] = {}
    for field, value in filters.items():
        norm_field = normalize_filter_field_name(field)
        if norm_field not in allowed:
            ignored[field] = value
            continue
        normalized[norm_field] = value
    if ignored:
        normalized["_ignoredFilters"] = ignored
    return normalized or None


def build_filter_payload(
    filters: Optional[Dict[str, Any]],
    min_price: Optional[float],
    max_price: Optional[float],
    price: Optional[float] = None,
) -> Optional[Dict[str, Any]]:
    payload: Dict[str, Any] = {}
    if filters:
        payload.update({k: v for k, v in filters.items() if not k.startswith("_")})
    if price is not None:
        payload["minPrice"] = float(price)
        payload["maxPrice"] = float(price)
    else:
        if min_price is not None:
            payload["minPrice"] = float(min_price)
        if max_price is not None:
            payload["maxPrice"] = float(max_price)
    return payload or None


def amount_as_int(row: Dict[str, Any]) -> int:
    try:
        return int(float(row.get("amount") or 0))
    except (TypeError, ValueError):
        return 0


def price_as_float(row: Dict[str, Any]) -> float:
    try:
        return float(row.get("price") or 0)
    except (TypeError, ValueError):
        return 0.0


def sum_qty_value(rows: List[Dict[str, Any]]) -> Tuple[int, float]:
    qty = sum(amount_as_int(row) for row in rows)
    value = sum(price_as_float(row) * amount_as_int(row) for row in rows)
    return qty, value


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "None", "nan"}:
        return None
    text = text.replace("\u00a0", "").replace(" ", "").replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def contact_lenses_totals(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    # For contact lenses we prefer explicit fields when present:
    # - qty_packs: "Количество, уп."
    # - qty_units: "Количество, шт."
    # - remaining_in_pack: "Осталось в уп."
    # - in_pack: "Кол-во в уп."
    # - value: "Сумма"
    # Otherwise we fall back to remoteRemains fields: amount/price and try to account for open packs
    # via remaining_in_pack + in_pack when present.
    total_packs = 0
    total_units = 0
    total_value = 0.0
    open_packs_count = 0

    for row in rows:
        qty_packs = parse_float(row.get("qty_packs") or row.get("Количество, уп.") or row.get("packs"))
        qty_units = parse_float(row.get("qty_units") or row.get("Количество, шт.") or row.get("units") or row.get("qty"))
        remaining_in_pack = parse_float(
            row.get("remaining_in_pack")
            or row.get("Осталось в уп.")
            or row.get("remainingInPack")
            or row.get("restInPack")
        )
        in_pack = parse_float(
            row.get("in_pack")
            or row.get("Кол-во в уп.")
            or row.get("inPack")
            or row.get("inpack")
        )
        value = parse_float(row.get("value") or row.get("Сумма") or row.get("sum") or row.get("total"))

        # 1) If report-style fields exist, use them as source of truth for this row.
        used_report_fields = False
        if qty_packs is not None or qty_units is not None or value is not None:
            used_report_fields = True
            total_packs += int(qty_packs or 0)
            total_units += int(qty_units or 0)
            total_value += float(value or 0)

        # 2) Otherwise fall back to remoteRemains.
        if not used_report_fields:
            packs = amount_as_int(row)
            price = price_as_float(row)
            total_packs += packs
            total_value += price * packs

            # If we have open-pack info, add 1 pack + remaining units for the open pack.
            if remaining_in_pack is not None and in_pack is not None:
                if remaining_in_pack > 0 and remaining_in_pack < in_pack:
                    open_packs_count += 1
                    total_packs += 1
                    total_units += int(remaining_in_pack)
                else:
                    # Best-effort: if only in_pack exists and no open pack, estimate units from full packs.
                    total_units += int(packs * in_pack)

    return {
        "total_qty_packs": total_packs,
        "total_qty_units": total_units,
        "total_value": round(total_value, 2),
        "open_packs_detected": open_packs_count,
    }


def normalize_header(text: str) -> str:
    t = str(text or "").strip().lower()
    t = t.replace("\u00a0", " ")
    while "  " in t:
        t = t.replace("  ", " ")
    t = t.strip(":;,.")
    return t


REMAINGOODS_HEADERS = {
    "department": {"департамент", "department", "салон", "магазин"},
    "qty_packs": {"количество, уп.", "количество уп.", "кол-во, уп.", "кол-во уп.", "количество упаковок", "упаковки"},
    "qty_units": {"количество, шт.", "количество шт.", "кол-во, шт.", "кол-во шт.", "количество штук", "штуки", "шт"},
    "remaining_in_pack": {"осталось в уп.", "осталось в уп", "остаток в уп.", "осталось"},
    "in_pack": {"кол-во в уп.", "кол-во в уп", "количество в уп.", "количество в уп", "в упаковке", "кол-во в упаковке"},
    # Important: ITigris export has both "Стоимость" (unit price) and "Сумма" (total).
    # For truth totals we must use "Сумма".
    "value": {"сумма", "сумма, руб", "сумма руб", "сумма (руб)"},
    "unit_price": {"стоимость", "цена", "цена, руб", "цена руб"},
}

REMAINGOODS_HEADERS_NORM: Dict[str, set[str]] = {
    key: {normalize_header(v) for v in variants} for key, variants in REMAINGOODS_HEADERS.items()
}


def find_col_index(headers: List[str], variants: set[str]) -> Optional[int]:
    for idx, header in enumerate(headers):
        if normalize_header(header) in variants:
            return idx
    return None


def detect_header_indices(headers: List[str]) -> Dict[str, Optional[int]]:
    return {
        "department": find_col_index(headers, REMAINGOODS_HEADERS_NORM["department"]),
        "qty_packs": find_col_index(headers, REMAINGOODS_HEADERS_NORM["qty_packs"]),
        "qty_units": find_col_index(headers, REMAINGOODS_HEADERS_NORM["qty_units"]),
        "remaining_in_pack": find_col_index(headers, REMAINGOODS_HEADERS_NORM["remaining_in_pack"]),
        "in_pack": find_col_index(headers, REMAINGOODS_HEADERS_NORM["in_pack"]),
        "value": find_col_index(headers, REMAINGOODS_HEADERS_NORM["value"]),
        "unit_price": find_col_index(headers, REMAINGOODS_HEADERS_NORM["unit_price"]),
    }


def header_indices_ok(idxs: Dict[str, Optional[int]]) -> bool:
    return idxs["qty_packs"] is not None and idxs["qty_units"] is not None and idxs["value"] is not None


def debug_headers(headers: List[str]) -> Dict[str, Any]:
    idxs = detect_header_indices(headers)
    found = {k: headers[v] if v is not None and v < len(headers) else None for k, v in idxs.items()}
    return {"indices": idxs, "found_headers": found, "normalized_headers": [normalize_header(h) for h in headers]}


def parse_remain_goods_csv(raw: bytes) -> List[Dict[str, Any]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        return []
    header_row_index: Optional[int] = None
    last_debug: Optional[Dict[str, Any]] = None
    for i in range(0, min(50, len(all_rows))):
        headers = [str(h or "") for h in all_rows[i]]
        dbg = debug_headers(headers)
        last_debug = dbg
        if header_indices_ok(dbg["indices"]):
            header_row_index = i
            break
    if header_row_index is None:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "remainGoodsReport_missing_required_columns",
                "hint": "Header row not detected in first 50 rows.",
                "preview_rows": all_rows[:20],
                "headers_debug": last_debug,
            },
        )

    headers = [str(h or "") for h in all_rows[header_row_index]]
    hdr_dbg = debug_headers(headers)
    idx_department = hdr_dbg["indices"]["department"]
    idx_packs = hdr_dbg["indices"]["qty_packs"]
    idx_units = hdr_dbg["indices"]["qty_units"]
    idx_remaining = hdr_dbg["indices"]["remaining_in_pack"]
    idx_in_pack = hdr_dbg["indices"]["in_pack"]
    idx_value = hdr_dbg["indices"]["value"]
    idx_unit_price = hdr_dbg["indices"]["unit_price"]

    out: List[Dict[str, Any]] = []
    for row in all_rows[header_row_index + 1 :]:
        def get_i(i: Optional[int]) -> Any:
            if i is None or i >= len(row):
                return None
            return row[i]

        out.append(
            {
                "department": str(get_i(idx_department) or "").strip() if idx_department is not None else "",
                "qty_packs": parse_float(get_i(idx_packs)) or 0,
                "qty_units": parse_float(get_i(idx_units)) or 0,
                "remaining_in_pack": parse_float(get_i(idx_remaining)) if idx_remaining is not None else None,
                "in_pack": parse_float(get_i(idx_in_pack)) if idx_in_pack is not None else None,
                "value": parse_float(get_i(idx_value)) or 0,
                "unit_price": parse_float(get_i(idx_unit_price)) if idx_unit_price is not None else None,
            }
        )
    return out


def parse_remain_goods_xlsx(raw: bytes) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    best_debug: Optional[Dict[str, Any]] = None
    best_preview: List[List[str]] = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        preview = []
        rows_cache: List[List[Any]] = []
        for r_i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = list(row)
            rows_cache.append(vals)
            if r_i < 20:
                preview.append([str(v or "") for v in vals[:25]])
            if r_i >= 60:
                break
        if not rows_cache:
            continue

        header_row_index: Optional[int] = None
        last_debug: Optional[Dict[str, Any]] = None
        for i in range(0, min(50, len(rows_cache))):
            headers = [str(h or "") for h in rows_cache[i]]
            dbg = debug_headers(headers)
            last_debug = dbg
            if header_indices_ok(dbg["indices"]):
                header_row_index = i
                break

        if header_row_index is None:
            if best_debug is None:
                best_debug = {"sheet": sheet_name, "headers_debug": last_debug}
                best_preview = preview
            continue

        headers = [str(h or "") for h in rows_cache[header_row_index]]
        hdr_dbg = debug_headers(headers)
        idx_department = hdr_dbg["indices"]["department"]
        idx_packs = hdr_dbg["indices"]["qty_packs"]
        idx_units = hdr_dbg["indices"]["qty_units"]
        idx_remaining = hdr_dbg["indices"]["remaining_in_pack"]
        idx_in_pack = hdr_dbg["indices"]["in_pack"]
        idx_value = hdr_dbg["indices"]["value"]
        idx_unit_price = hdr_dbg["indices"]["unit_price"]

        out: List[Dict[str, Any]] = []
        for row in rows_cache[header_row_index + 1 :]:
            def get_i(i: Optional[int]) -> Any:
                if i is None or i >= len(row):
                    return None
                return row[i]

            out.append(
                {
                    "department": str(get_i(idx_department) or "").strip() if idx_department is not None else "",
                    "qty_packs": parse_float(get_i(idx_packs)) or 0,
                    "qty_units": parse_float(get_i(idx_units)) or 0,
                    "remaining_in_pack": parse_float(get_i(idx_remaining)) if idx_remaining is not None else None,
                    "in_pack": parse_float(get_i(idx_in_pack)) if idx_in_pack is not None else None,
                    "value": parse_float(get_i(idx_value)) or 0,
                    "unit_price": parse_float(get_i(idx_unit_price)) if idx_unit_price is not None else None,
                }
            )
        return out

    raise HTTPException(
        status_code=400,
        detail={
            "error": "remainGoodsReport_missing_required_columns",
            "sheets": sheet_names,
            "preview_rows": best_preview,
            "debug": best_debug,
        },
    )


def parse_remain_goods_xls(raw: bytes) -> List[Dict[str, Any]]:
    book = xlrd.open_workbook(file_contents=raw)
    sheet_names = book.sheet_names()
    best_debug: Optional[Dict[str, Any]] = None
    best_preview: List[List[str]] = []

    for s_i, sheet_name in enumerate(sheet_names):
        sheet = book.sheet_by_index(s_i)
        if sheet.nrows <= 0:
            continue
        preview = [
            [str(sheet.cell_value(r, c) or "") for c in range(min(sheet.ncols, 25))]
            for r in range(min(sheet.nrows, 20))
        ]

        header_row_index: Optional[int] = None
        last_debug: Optional[Dict[str, Any]] = None
        for r in range(0, min(50, sheet.nrows)):
            headers = [str(sheet.cell_value(r, c) or "") for c in range(sheet.ncols)]
            dbg = debug_headers(headers)
            last_debug = dbg
            if header_indices_ok(dbg["indices"]):
                header_row_index = r
                break

        if header_row_index is None:
            if best_debug is None:
                best_debug = {"sheet": sheet_name, "headers_debug": last_debug}
                best_preview = preview
            continue

        headers = [str(sheet.cell_value(header_row_index, c) or "") for c in range(sheet.ncols)]
        hdr_dbg = debug_headers(headers)
        idx_department = hdr_dbg["indices"]["department"]
        idx_packs = hdr_dbg["indices"]["qty_packs"]
        idx_units = hdr_dbg["indices"]["qty_units"]
        idx_remaining = hdr_dbg["indices"]["remaining_in_pack"]
        idx_in_pack = hdr_dbg["indices"]["in_pack"]
        idx_value = hdr_dbg["indices"]["value"]
        idx_unit_price = hdr_dbg["indices"]["unit_price"]

        def cell(r: int, c: Optional[int]) -> Any:
            if c is None or c < 0 or c >= sheet.ncols:
                return None
            if r < 0 or r >= sheet.nrows:
                return None
            return sheet.cell_value(r, c)

        out: List[Dict[str, Any]] = []
        for r in range(header_row_index + 1, sheet.nrows):
            out.append(
                {
                    "department": str(cell(r, idx_department) or "").strip() if idx_department is not None else "",
                    "qty_packs": parse_float(cell(r, idx_packs)) or 0,
                    "qty_units": parse_float(cell(r, idx_units)) or 0,
                    "remaining_in_pack": parse_float(cell(r, idx_remaining)) if idx_remaining is not None else None,
                    "in_pack": parse_float(cell(r, idx_in_pack)) if idx_in_pack is not None else None,
                    "value": parse_float(cell(r, idx_value)) or 0,
                    "unit_price": parse_float(cell(r, idx_unit_price)) if idx_unit_price is not None else None,
                }
            )
        return out

    raise HTTPException(
        status_code=400,
        detail={
            "error": "remainGoodsReport_missing_required_columns",
            "sheets": sheet_names,
            "preview_rows": best_preview,
            "debug": best_debug,
        },
    )


def remain_goods_totals(rows: List[Dict[str, Any]], prefer_summary_row: bool = True) -> Dict[str, Any]:
    # Prefer explicit summary line if present (common in ITigris exports).
    summary_row: Optional[Dict[str, Any]] = None
    if prefer_summary_row:
        for r in rows:
            dep = normalize_header(r.get("department") or "")
            if dep in {"итого", "всего", "итог", "total"}:
                summary_row = r
                break

    value_from_details = False
    if summary_row:
        total_packs = int(parse_float(summary_row.get("qty_packs")) or 0)
        total_units = int(parse_float(summary_row.get("qty_units")) or 0)
        total_value = float(parse_float(summary_row.get("value")) or 0)
        # Some .xls exports keep totals as formulas; xlrd may read them as 0.
        if total_value <= 0:
            detail_value = sum(
                (parse_float(r.get("value")) or 0)
                for r in rows
                if normalize_header(r.get("department") or "") not in {"итого", "всего", "итог", "total"}
            )
            if detail_value > 0:
                total_value = float(detail_value)
                value_from_details = True
    else:
        total_packs = int(sum(parse_float(r.get("qty_packs")) or 0 for r in rows))
        total_units = int(sum(parse_float(r.get("qty_units")) or 0 for r in rows))
        total_value = float(sum(parse_float(r.get("value")) or 0 for r in rows))
    open_lines = 0
    open_packs = 0
    open_units = 0
    open_value = 0.0
    for r in rows:
        rem = parse_float(r.get("remaining_in_pack"))
        in_pack = parse_float(r.get("in_pack"))
        if rem is None or in_pack is None:
            continue
        if rem > 0 and rem < in_pack:
            open_lines += 1
            open_packs += int(parse_float(r.get("qty_packs")) or 0)
            open_units += int(parse_float(r.get("qty_units")) or 0)
            open_value += float(parse_float(r.get("value")) or 0)

    return {
        "total_qty_packs": total_packs,
        "total_qty_units": total_units,
        "total_value": round(total_value, 2),
        "open_lines": open_lines,
        "open_qty_packs": open_packs,
        "open_qty_units": open_units,
        "open_value": round(open_value, 2),
        "used_summary_row": bool(summary_row),
        "value_from_details_sum": value_from_details,
    }


def parse_remain_goods_by_filename(filename: str, raw: bytes) -> List[Dict[str, Any]]:
    low = (filename or "").lower()
    if low.endswith(".csv"):
        return parse_remain_goods_csv(raw)
    if low.endswith(".xlsx"):
        return parse_remain_goods_xlsx(raw)
    if low.endswith(".xls"):
        return parse_remain_goods_xls(raw)
    raise HTTPException(status_code=400, detail="unsupported_report_file_type")


def get_snapshot(department_id: int) -> Optional[Dict[str, Any]]:
    snap = _contactlenses_report_snapshots.get(department_id)
    if not snap:
        return None
    if int(time.time()) >= int(snap["expires_at_unix"]):
        _contactlenses_report_snapshots.pop(department_id, None)
        return None
    return snap


def get_global_snapshot() -> Optional[Dict[str, Any]]:
    global _contactlenses_report_global_snapshot
    snap = _contactlenses_report_global_snapshot
    if not snap:
        return None
    if int(time.time()) >= int(snap["expires_at_unix"]):
        _contactlenses_report_global_snapshot = None
        return None
    return snap


def _render_itigris_url(template: str) -> str:
    # Supported placeholders: {app}, {key}, {external_key}
    return template.format(app=APP_NAME, key=REMOTE_API_KEY, external_key=EXTERNAL_API_KEY)


def _guess_filename_from_content_disposition(headers: httpx.Headers) -> Optional[str]:
    cd = headers.get("content-disposition") or headers.get("Content-Disposition")
    if not cd:
        return None
    m = re.search(r"filename\\*=UTF-8''([^;]+)", cd, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r'filename=\"?([^\";]+)\"?', cd, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return None


async def _auto_fetch_remain_goods_report_bytes() -> Tuple[Optional[str], Optional[bytes], Optional[str]]:
    """
    Returns: (filename, raw_bytes, error_code)
    """
    if not REMAINGOODS_AUTO_FETCH_URL_TEMPLATE:
        return None, None, "auto_fetch_not_configured"
    if not REMOTE_API_KEY and ("{key}" in REMAINGOODS_AUTO_FETCH_URL_TEMPLATE):
        return None, None, "remote_api_key_missing"

    url = _render_itigris_url(REMAINGOODS_AUTO_FETCH_URL_TEMPLATE)
    try:
        async with httpx.AsyncClient(timeout=TIMEOUT, follow_redirects=True) as client:
            resp = await client.get(url)
    except Exception:
        return None, None, "auto_fetch_network_error"

    if resp.status_code != 200:
        return None, None, f"auto_fetch_http_{resp.status_code}"

    raw = resp.content
    if not raw:
        return None, None, "auto_fetch_empty_response"
    if len(raw) > 30_000_000:
        return None, None, "auto_fetch_file_too_large"

    filename = _guess_filename_from_content_disposition(resp.headers)
    return filename, raw, None


def _parse_remain_goods_html(raw: bytes) -> List[Dict[str, Any]]:
    """
    Parse remainGoodsReport HTML page and convert to normalized rows.
    HTML layout can change; we search for a table that contains required headers.
    """
    html = raw.decode("utf-8", errors="replace")
    soup = BeautifulSoup(html, "html.parser")

    required_keys = {"qty_packs", "qty_units", "remaining_in_pack", "in_pack", "value"}
    tables = soup.find_all("table")
    best: Optional[Tuple[List[str], Any]] = None
    debug_tables: List[Dict[str, Any]] = []
    for table in tables:
        # Find first row that looks like header.
        rows = table.find_all("tr")
        if not rows:
            continue
        header_row = None
        header_cells: List[str] = []
        candidates: List[Dict[str, Any]] = []
        for tr in rows[:30]:
            cells = tr.find_all(["th", "td"])
            if not cells:
                continue
            texts = [c.get_text(" ", strip=True) for c in cells]
            norms = [normalize_header(t) for t in texts]
            idx = {
                "department": find_col_index(norms, REMAINGOODS_HEADERS_NORM["department"]),
                "qty_packs": find_col_index(norms, REMAINGOODS_HEADERS_NORM["qty_packs"]),
                "qty_units": find_col_index(norms, REMAINGOODS_HEADERS_NORM["qty_units"]),
                "remaining_in_pack": find_col_index(norms, REMAINGOODS_HEADERS_NORM["remaining_in_pack"]),
                "in_pack": find_col_index(norms, REMAINGOODS_HEADERS_NORM["in_pack"]),
                "value": find_col_index(norms, REMAINGOODS_HEADERS_NORM["value"]),
                "unit_price": find_col_index(norms, REMAINGOODS_HEADERS_NORM["unit_price"]),
            }
            present = {k for k, v in idx.items() if v is not None}
            candidates.append({"cells": texts[:20], "present_keys": sorted(present)})
            if required_keys.issubset(present):
                header_row = tr
                header_cells = texts
                break
        if not header_row:
            if candidates:
                debug_tables.append({"candidates": candidates[:5]})
            continue
        best = (header_cells, table)
        break

    if not best:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "remainGoodsReport_missing_required_columns",
                "tables_found": len(tables),
                "debug_tables": debug_tables[:3],
            },
        )

    header_cells, table = best
    headers = [normalize_header(h) for h in header_cells]
    idx = {
        "department": find_col_index(headers, REMAINGOODS_HEADERS_NORM["department"]),
        "qty_packs": find_col_index(headers, REMAINGOODS_HEADERS_NORM["qty_packs"]),
        "qty_units": find_col_index(headers, REMAINGOODS_HEADERS_NORM["qty_units"]),
        "remaining_in_pack": find_col_index(headers, REMAINGOODS_HEADERS_NORM["remaining_in_pack"]),
        "in_pack": find_col_index(headers, REMAINGOODS_HEADERS_NORM["in_pack"]),
        "value": find_col_index(headers, REMAINGOODS_HEADERS_NORM["value"]),
        "unit_price": find_col_index(headers, REMAINGOODS_HEADERS_NORM["unit_price"]),
    }

    def cell_text(c: Any) -> str:
        return (c.get_text(" ", strip=True) if c else "").strip()

    def parse_num(s: str) -> float:
        # Supports "2 320 150,00" and "2320150.00" formats.
        t = (s or "").strip()
        if not t:
            return 0.0
        t = t.replace("\xa0", " ").replace(" ", "")
        t = t.replace("руб.", "").replace("руб", "").replace("₽", "")
        t = t.replace(",", ".")
        # keep digits, minus, dot
        t = re.sub(r"[^0-9\\-\\.]", "", t)
        if t in {"", "-", "."}:
            return 0.0
        try:
            return float(t)
        except Exception:
            return 0.0

    rows_out: List[Dict[str, Any]] = []
    for tr in table.find_all("tr"):
        cells = tr.find_all(["td", "th"])
        if not cells:
            continue
        texts = [cell_text(c) for c in cells]
        # skip obvious header-like rows
        if any(normalize_header(t) == headers[0] for t in texts[:1]) and len(texts) == len(headers):
            continue
        if len(texts) < len(headers):
            continue
        dep = texts[idx["department"]] if idx["department"] is not None else ""
        qty_packs = parse_num(texts[idx["qty_packs"]]) if idx["qty_packs"] is not None else 0.0
        qty_units = parse_num(texts[idx["qty_units"]]) if idx["qty_units"] is not None else 0.0
        remaining_in_pack = parse_num(texts[idx["remaining_in_pack"]]) if idx["remaining_in_pack"] is not None else 0.0
        in_pack = parse_num(texts[idx["in_pack"]]) if idx["in_pack"] is not None else 0.0
        value = parse_num(texts[idx["value"]]) if idx["value"] is not None else 0.0
        unit_price = parse_num(texts[idx["unit_price"]]) if idx["unit_price"] is not None else 0.0

        if not dep and qty_packs == 0 and qty_units == 0 and value == 0:
            continue

        rows_out.append(
            {
                "department": dep,
                "qty_packs": qty_packs,
                "qty_units": qty_units,
                "remaining_in_pack": remaining_in_pack,
                "in_pack": in_pack,
                "value": value,
                "unit_price": unit_price,
            }
        )
    if not rows_out:
        raise HTTPException(status_code=400, detail="remainGoodsReport_no_rows_detected")
    return rows_out


def _extract_optima_ctx_from_text(text: str) -> Dict[str, str]:
    """
    Best-effort extractor for userId/pageUUID/uuidValue/companyUUID from Optima HTML/JS responses.
    Does not assume a specific template.
    """
    src = text or ""

    def _find_hidden(name: str) -> str:
        m = re.search(rf'name=\"{re.escape(name)}\"\\s+value=\"([^\"]+)\"', src, flags=re.IGNORECASE)
        return m.group(1).strip() if m else ""

    def _find_query(name: str) -> str:
        # Match ...name=VALUE... in text blobs (hrefs/scripts).
        m = re.search(rf"(?:\\?|&){re.escape(name)}=([^&\\s\\\"'>]+)", src, flags=re.IGNORECASE)
        return m.group(1).strip() if m else ""

    def _find_js(name: str) -> str:
        # Match name:'value' or name=\"value\"
        m = re.search(rf"{re.escape(name)}\\s*[:=]\\s*['\\\"]([^'\\\"]+)['\\\"]", src, flags=re.IGNORECASE)
        return m.group(1).strip() if m else ""

    def _find_anywhere(name: str) -> str:
        # More permissive: pageUUID\\u003d<uuid>, pageUUID%3D<uuid>, pageUUID=<uuid>
        m = re.search(rf"{re.escape(name)}(?:\\\\u003d|%3[Dd]|=)\\s*({uuid_re.pattern})", src, flags=re.IGNORECASE)
        return m.group(1).strip() if m else ""

    uuid_re = re.compile(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}")

    def _collect_uuids(limit: int = 50) -> List[str]:
        seen: List[str] = []
        for m in uuid_re.finditer(src):
            u = m.group(0)
            if u not in seen:
                seen.append(u)
            if len(seen) >= limit:
                break
        return seen

    def _find_keyed_uuid(keys: List[str]) -> str:
        # Look for "<key> ... <uuid>" patterns in HTML/JS.
        for k in keys:
            m = re.search(rf"{re.escape(k)}[^0-9a-fA-F]{{0,40}}({uuid_re.pattern})", src, flags=re.IGNORECASE)
            if m:
                return m.group(1)
        return ""

    out: Dict[str, str] = {}

    def _safe_search(pattern: str) -> Optional[re.Match]:
        try:
            return re.search(pattern, src, flags=re.IGNORECASE)
        except re.error:
            # Never fail the whole request due to a heuristic extractor bug.
            return None

    # userId is numeric, not uuid.
    out["userId"] = _find_hidden("userId") or _find_query("userId") or _find_js("userId") or ""

    # pageUUID / uuidValue may appear with different casing.
    out["pageUUID"] = (
        _find_hidden("pageUUID")
        or _find_hidden("pageUuid")
        or _find_query("pageUUID")
        or _find_query("pageUuid")
        or _find_js("pageUUID")
        or _find_js("pageUuid")
        or _find_anywhere("pageUUID")
        or _find_anywhere("pageUuid")
        or _find_keyed_uuid(["pageUUID", "pageUuid", "page_uuid"])
        or ""
    )
    out["uuidValue"] = (
        _find_hidden("uuidValue")
        or _find_hidden("uuidvalue")
        or _find_query("uuidValue")
        or _find_query("uuidvalue")
        or _find_js("uuidValue")
        or _find_js("uuidvalue")
        or _find_anywhere("uuidValue")
        or _find_anywhere("uuidvalue")
        or _find_keyed_uuid(["uuidValue", "uuidvalue", "uuid_value"])
        or ""
    )
    out["companyUUID"] = (
        _find_hidden("companyUUID")
        or _find_hidden("companyUuid")
        or _find_query("companyUUID")
        or _find_query("companyUuid")
        or _find_js("companyUUID")
        or _find_js("companyUuid")
        or ""
    )

    # Heuristic: some pages carry page context in JS function calls, e.g. updateMainPageData(..., "<uuid>", ...).
    if not out["pageUUID"]:
        # Match both "updateMainPageData(" and "updateMainPageData\\(" (some pages escape "(" as "\\(" in JS strings).
        m = _safe_search(rf"updateMainPageData\\?\\([^)]*?['\\\"]({uuid_re.pattern})['\\\"]")
        if m:
            out["pageUUID"] = m.group(1)

    # Another common variant: updateMainPageDataMin("<pageUUID>")
    if not out["pageUUID"]:
        m = _safe_search(rf"updateMainPageDataMin\\?\\(['\\\"]({uuid_re.pattern})['\\\"]\\)")
        if m:
            out["pageUUID"] = m.group(1)

    # remainGoodsReport: startPage HTML often carries the *next* uuidValue (used for reportPage/prepareData)
    # inside updateMainPageData(..., "<uuidValue>", ""), without an explicit "uuidValue=" label.
    if not out["uuidValue"]:
        m = _safe_search(rf"updateMainPageData\\?\\([^)]*?['\\\"]({uuid_re.pattern})['\\\"]\\s*,\\s*['\\\"]\\s*['\\\"]\\s*\\)")
        if m:
            out["uuidValue"] = m.group(1)

    # Attach uuid candidates for debugging when extraction fails.
    if not out["pageUUID"] or not out["uuidValue"]:
        out["_uuid_candidates"] = ",".join(_collect_uuids(25))
    return out


async def _auto_fetch_remain_goods_report_via_web(date_ddmmyyyy: Optional[str] = None) -> Tuple[str, List[Dict[str, Any]]]:
    """
    Uses Optima web session to request HTML reportPage and parse it.
    Preferred: login/password/key (server performs login to get fresh session cookies).
    Fallback: static Cookie header (ITIGRIS_REMAINGOODSREPORT_WEB_COOKIE).
    """
    company_uuid = REMAINGOODS_WEB_COMPANY_UUID or APP_NAME

    if not date_ddmmyyyy:
        # Default: server local date (Render is usually UTC; for MVP this is OK, can override via refresh endpoint).
        date_ddmmyyyy = time.strftime("%d.%m.%Y", time.localtime())

    if REMAINGOODS_WEB_DEPARTMENT_IDS:
        dep_ids = [d.strip() for d in REMAINGOODS_WEB_DEPARTMENT_IDS.split(",") if d.strip()]
    else:
        dep_ids = [str(v) for v in DEPARTMENTS.values()]

    def build_report_form(user_id: str, page_uuid: str, uuid_value: str) -> Dict[str, Any]:
        # NOTE: For httpx.AsyncClient we must avoid passing "data" as a list of tuples with repeated
        # keys, because httpx may build a sync byte stream (IteratorByteStream) which fails at send time.
        # Use dict + list values instead.
        # NOTE: Optima's report endpoints may expect the full form payload (even if fields are empty).
        # We mirror the browser payload shape, defaulting optional filters to empty strings.
        return {
        "date": date_ddmmyyyy,
        # department selection
        "department_input0": "",
        "department": dep_ids,
        # supply/supplier/price ranges
        "supplyDateFrom": "",
        "supplyDateTo": "",
        "supplier_input0": "",
        "supplier": "",
        "priceFrom": "",
        "priceTo": "",
        # grouping
        "groupByDepartment_input0": "",
        "groupByDepartment": "true",
        # report type / price type
        "reportType_input0": REMAINGOODS_WEB_REPORT_TYPE,
        "reportType": REMAINGOODS_WEB_REPORT_TYPE,
        "priceType_input0": REMAINGOODS_WEB_PRICE_TYPE,
        "priceType": REMAINGOODS_WEB_PRICE_TYPE,
        # accessory filters
        "accessory.category_input0": "",
        "accessory.category": "",
        "accessory.model": "",
        # contact lenses filters
        "cl.manufacturer_input0": "",
        "cl.manufacturer": "",
        "cl.name": "",
        "cl.packageNum_input0": "",
        "cl.packageNum": "",
        "cl.color_input0": "",
        "cl.color": "",
        "cl.dioptresFrom_input0": "",
        "cl.dioptresFrom": "",
        "cl.dioptresTo_input0": "",
        "cl.dioptresTo": "",
        "cl.cylinderFrom_input0": "",
        "cl.cylinderFrom": "",
        "cl.cylinderTo_input0": "",
        "cl.cylinderTo": "",
        "cl.radiusOfCurvature_input0": "",
        "cl.radiusOfCurvature": "",
        "cl.diameter_input0": "",
        "cl.diameter": "",
        "cl.expirationDateFrom": "",
        "cl.expirationDateTo": "",
        "cl.wearingPeriod_input0": "",
        "cl.wearingPeriod": "",
        "cl._groupByExpirationDate": "",
        "cl._criticalDateOnly": "",
        # lenses filters
        "lenses.manufacturer_input0": "",
        "lenses.manufacturer": "",
        "lenses.brand_input0": "",
        "lenses.brand": "",
        "lenses.index_input0": "",
        "lenses.index": "",
        "lenses.diameter_input0": "",
        "lenses.diameter": "",
        "lenses.type_input0": "",
        "lenses.type": "",
        "lenses.material_input0": "",
        "lenses.material": "",
        "lenses.dioptresFrom_input0": "",
        "lenses.dioptresFrom": "",
        "lenses.dioptresTo_input0": "",
        "lenses.dioptresTo": "",
        "lenses.cylinderFrom_input0": "",
        "lenses.cylinderFrom": "",
        "lenses.cylinderTo_input0": "",
        "lenses.cylinderTo": "",
        "lenses.geometry_input0": "",
        "lenses.geometry": "",
        "lenses.lensClass_input0": "",
        "lenses.lensClass": "",
        "lenses.technology_input0": "",
        "lenses.technology": "",
        "lenses.color_input0": "",
        "lenses.color": "",
        "lenses.cover_input0": "",
        "lenses.cover": "",
        # glasses filters
        "glasses.manufacturer_input0": "",
        "glasses.manufacturer": "",
        "glasses.brand_input0": "",
        "glasses.brand": "",
        "glasses.model": "",
        "glasses.material_input0": "",
        "glasses.material": "",
        "glasses.type_input0": "",
        "glasses.type": "",
        "glasses.targetGroup_input0": "",
        "glasses.targetGroup": "",
        "glasses.design_input0": "",
        "glasses.design": "",
        # extras/glasses
        "extras-glasses.manufacturer_input0": "",
        "extras-glasses.manufacturer": "",
        "extras-glasses.brand_input0": "",
        "extras-glasses.brand": "",
        "extras-glasses.model": "",
        "extras-glasses.material_input0": "",
        "extras-glasses.material": "",
        "extras-glasses.type_input0": "",
        "extras-glasses.type": "",
        "extras-glasses.targetGroup_input0": "",
        "extras-glasses.targetGroup": "",
        # extras/lenses
        "extras-lenses.dioptresFrom_input0": "",
        "extras-lenses.dioptresFrom": "",
        "extras-lenses.dioptresTo_input0": "",
        "extras-lenses.dioptresTo": "",
        "extras-lenses.cylinderFrom_input0": "",
        "extras-lenses.cylinderFrom": "",
        "extras-lenses.cylinderTo_input0": "",
        "extras-lenses.cylinderTo": "",
        "extras-lenses.manufacturer_input0": "",
        "extras-lenses.manufacturer": "",
        "extras-lenses.brand_input0": "",
        "extras-lenses.brand": "",
        # semi-finished
        "semi.manufacturer_input0": "",
        "semi.manufacturer": "",
        "semi.brand_input0": "",
        "semi.brand": "",
        "semi.index_input0": "",
        "semi.index": "",
        "semi.diameter_input0": "",
        "semi.diameter": "",
        "semi.base_input0": "",
        "semi.base": "",
        "semi.material_input0": "",
        "semi.material": "",
        "semi.geometry_input0": "",
        "semi.geometry": "",
        "semi.type_input0": "",
        "semi.type": "",
        "semi.technology_input0": "",
        "semi.technology": "",
        "semi.color_input0": "",
        "semi.color": "",
        "semi.lensClass_input0": "",
        "semi.lensClass": "",
        "semi.cover_input0": "",
        "semi.cover": "",
        # sunglasses
        "sunglasses.manufacturer_input0": "",
        "sunglasses.manufacturer": "",
        "sunglasses.brand_input0": "",
        "sunglasses.brand": "",
        "sunglasses.model": "",
        "sunglasses.material_input0": "",
        "sunglasses.material": "",
        "sunglasses.type_input0": "",
        "sunglasses.type": "",
        "sunglasses.frameType_input0": "",
        "sunglasses.frameType": "",
        "sunglasses.targetGroup_input0": "",
        "sunglasses.targetGroup": "",
        "sunglasses.design_input0": "",
        "sunglasses.design": "",
        # auth/session context
        "userId": user_id,
        "uuidValue": uuid_value,
        "pageUUID": page_uuid,
        "companyUUID": company_uuid,
        }

    url = f"https://optima.itigris.ru/{APP_NAME}/remainGoodsReport/reportPage"
    ua = ITIGRIS_WEB_USER_AGENT or "Mozilla/5.0"
    headers = {"X-Requested-With": "XMLHttpRequest"}
    # Match browser-ish headers for XHR.
    base_url = f"https://optima.itigris.ru/{APP_NAME}"
    web_headers = {
        "X-Requested-With": "XMLHttpRequest",
        "User-Agent": ua,
        "Accept": "text/html, */*; q=0.01",
        "Referer": f"{base_url}/remainGoodsReport",
        "Origin": "https://optima.itigris.ru",
    }

    async def do_report_request(client: httpx.AsyncClient, user_id: str, page_uuid: str, uuid_value: str) -> httpx.Response:
        form = build_report_form(user_id=user_id, page_uuid=page_uuid, uuid_value=uuid_value)
        return await client.post(url, data=form, headers=headers)

    async def get_report_module_context(client: httpx.AsyncClient, user_id: str) -> Dict[str, str]:
        """
        Optima uses per-module pageUUID/uuidValue. For remainGoodsReport, browser initializes a separate
        page context before calling startPage/reportPage. We best-effort fetch module page and extract
        pageUUID/uuidValue from the HTML/JS.
        """
        candidates = [
            f"https://optima.itigris.ru/{APP_NAME}/remainGoodsReport",
            f"https://optima.itigris.ru/{APP_NAME}/remainGoodsReport/",
            f"https://optima.itigris.ru/{APP_NAME}/remainGoodsReport/startPage",
        ]
        for u in candidates:
            try:
                r = await client.get(u, headers=web_headers)
            except Exception:
                continue
            if r.status_code != 200:
                continue
            ctx = _extract_optima_ctx_from_text(r.text or "")
            if ctx.get("pageUUID") and ctx.get("uuidValue"):
                # Some pages don't include userId; keep provided one.
                ctx["userId"] = ctx.get("userId") or user_id
                return ctx
        return {}

    async def login_and_get_context(client: httpx.AsyncClient) -> Dict[str, str]:
        if not ITIGRIS_WEB_LOGIN or not ITIGRIS_WEB_PASSWORD or not ITIGRIS_WEB_KEY:
            raise HTTPException(status_code=500, detail="auto_web_login_not_configured")

        # Bootstrap cookies (JSESSIONID, route, etc.) and try to extract pageUUID/uuidValue if present.
        # In a real browser a GET happens before login.
        try:
            await client.get(f"https://optima.itigris.ru/{APP_NAME}", headers=web_headers)
        except Exception:
            pass
        login_page_text = ""
        try:
            r = await client.get(f"https://optima.itigris.ru/{APP_NAME}/login", headers=web_headers)
            login_page_text = r.text or ""
        except Exception:
            pass

        extracted_page_uuid = ""
        extracted_uuid_value = ""
        extracted_user_id = ""
        if login_page_text:
            extracted = _extract_optima_ctx_from_text(login_page_text)
            extracted_page_uuid = extracted.get("pageUUID") or ""
            extracted_uuid_value = extracted.get("uuidValue") or ""
            extracted_user_id = extracted.get("userId") or ""

        # If we can't extract from HTML, allow explicit env overrides.
        extracted_page_uuid = extracted_page_uuid or ITIGRIS_WEB_PAGE_UUID
        extracted_uuid_value = extracted_uuid_value or ITIGRIS_WEB_UUID_VALUE
        extracted_user_id = extracted_user_id or ITIGRIS_WEB_USER_ID

        login_form: Dict[str, Any] = {
            "loginAction": "true",
            "login": ITIGRIS_WEB_LOGIN,
            "password": ITIGRIS_WEB_PASSWORD,
            "key": ITIGRIS_WEB_KEY,
            "versionDesc": ITIGRIS_WEB_VERSION_DESC or "server",
            "browserDesc": ITIGRIS_WEB_BROWSER_DESC or ua,
            "userId": extracted_user_id or "",
            "uuidValue": extracted_uuid_value or "",
            "pageUUID": extracted_page_uuid or "",
            "companyUUID": company_uuid,
        }

        # We want to capture redirect params from Location.
        try:
            resp = await client.post(ITIGRIS_WEB_LOGIN_URL, data=login_form, headers=web_headers, follow_redirects=False)
        except TypeError:
            # Older httpx: follow_redirects is client-level only.
            resp = await client.post(ITIGRIS_WEB_LOGIN_URL, data=login_form, headers=web_headers)

        login_status = resp.status_code
        login_location = resp.headers.get("location") or resp.headers.get("Location") or ""
        login_set_cookie_present = bool(resp.headers.get("set-cookie") or resp.headers.get("Set-Cookie"))

        if login_status not in {200, 302, 303}:
            raise HTTPException(
                status_code=502,
                detail={
                    "error": "auto_web_login_failed",
                    "status_code": login_status,
                    "location": login_location or None,
                    "set_cookie_present": login_set_cookie_present,
                    "sent_payload_meta": {
                        "keys": sorted(list(login_form.keys())),
                        "companyUUID": login_form.get("companyUUID"),
                        "pageUUID_present": bool(login_form.get("pageUUID")),
                        "uuidValue_present": bool(login_form.get("uuidValue")),
                        "userId_present": bool(login_form.get("userId")),
                        "versionDesc": login_form.get("versionDesc"),
                        "browserDesc_present": bool(login_form.get("browserDesc")),
                    },
                    "body_snippet": (resp.text or "")[:1000],
                },
            )

        # If login returned 200 without redirect, it's almost always "bad credentials" or a blocked login page.
        # We consider it a failure, unless the caller explicitly relies on env context (legacy fallback).
        if login_status == 200 and not login_location:
            raise HTTPException(
                status_code=502,
                detail={
                    "error": "auto_web_login_no_redirect",
                    "status_code": login_status,
                    "set_cookie_present": login_set_cookie_present,
                    "body_snippet": (resp.text or "")[:1000],
                    "sent_payload_meta": {
                        "keys": sorted(list(login_form.keys())),
                        "companyUUID": login_form.get("companyUUID"),
                        "pageUUID_present": bool(login_form.get("pageUUID")),
                        "uuidValue_present": bool(login_form.get("uuidValue")),
                        "userId_present": bool(login_form.get("userId")),
                        "versionDesc": login_form.get("versionDesc"),
                        "browserDesc_present": bool(login_form.get("browserDesc")),
                    },
                },
            )

        location = login_location
        if location:
            absolute = urljoin(f"https://optima.itigris.ru/{APP_NAME}/", location.lstrip("/"))
            parsed = urlparse(absolute)
            qs = parse_qs(parsed.query)
            ctx = {
                "userId": (qs.get("userId") or [""])[0],
                "pageUUID": (qs.get("pageUUID") or [""])[0],
                "uuidValue": (qs.get("uuidValue") or [""])[0],
                "companyUUID": (qs.get("companyUUID") or [company_uuid])[0] or company_uuid,
            }
            # Hit userStart once; some sessions set additional cookies.
            user_start_ok = False
            try:
                r2 = await client.get(absolute, headers=web_headers)
                user_start_ok = (r2.status_code == 200)
            except Exception:
                user_start_ok = False

            # After login/userStart, Optima often issues the "current page context" (pageUUID/uuidValue)
            # on the main app page. remainGoodsReport/startPage expects those values in the POST payload.
            try:
                main_r = await client.get(f"https://optima.itigris.ru/{APP_NAME}", headers=web_headers)
                main_ctx = _extract_optima_ctx_from_text(main_r.text or "")
                # Capture extractor diagnostics for debugging (snippets are limited).
                main_text = main_r.text or ""
                # best-effort: find nearby snippet for updateMainPageDataMin/updateMainPageData
                snippet = ""
                for pat in ["updateMainPageDataMin", "updateMainPageData"]:
                    idx = main_text.find(pat)
                    if idx != -1:
                        start = max(0, idx - 120)
                        end = min(len(main_text), idx + 240)
                        snippet = main_text[start:end]
                        break
                # Keep userId from redirect, but refresh pageUUID/uuidValue from the main page when present.
                # Keep userId from redirect, but refresh pageUUID/uuidValue from the main page when present.
                if main_ctx.get("pageUUID"):
                    ctx["pageUUID"] = main_ctx.get("pageUUID") or ctx["pageUUID"]
                if main_ctx.get("uuidValue"):
                    ctx["uuidValue"] = main_ctx.get("uuidValue") or ctx["uuidValue"]
                # Attach debug (do not include full HTML).
                try:
                    dbg = ctx.get("_debug") or {}
                    dbg["main_page_probe"] = {
                        "status": main_r.status_code,
                        "len": len(main_text),
                        "extracted": {
                            "pageUUID": (main_ctx.get("pageUUID") or None),
                            "uuidValue": (main_ctx.get("uuidValue") or None),
                            "userId": (main_ctx.get("userId") or None),
                        },
                        "snippet": snippet or None,
                    }
                    ctx["_debug"] = dbg  # type: ignore[typeddict-item]
                except Exception:
                    pass
            except Exception:
                pass

            # include debug meta for downstream errors
            cookie_names = sorted({c.name for c in client.cookies.jar})
            ctx["_debug"] = {  # type: ignore[typeddict-item]
                "login_status": login_status,
                "login_location": location,
                "userStart_ok": user_start_ok,
                "cookie_names": cookie_names,
            }
            return ctx

        # No redirect location: this is unexpected at this point.
        cookie_names = sorted({c.name for c in client.cookies.jar})
        raise HTTPException(
            status_code=502,
            detail={
                "error": "auto_web_login_missing_location",
                "status_code": login_status,
                "set_cookie_present": login_set_cookie_present,
                "cookie_names": cookie_names,
                "body_snippet": (resp.text or "")[:1000],
            },
        )

    async def _follow_location(client: httpx.AsyncClient, loc: str) -> httpx.Response:
        absolute = urljoin(f"https://optima.itigris.ru/{APP_NAME}/", loc.lstrip("/"))
        return await client.get(absolute, headers=web_headers, follow_redirects=False)

    async def _init_accountant_reports_ctx(client: httpx.AsyncClient, base_ctx: Dict[str, str]) -> Dict[str, str]:
        """
        Reproduce the exact navigation chain seen in HAR before remainGoodsReport/startPage:
          POST /menu/openMenuElement?menuItem=accountantReports
          GET  /menu/openMenuElement2?menuItem=accountantReports&...
          GET  /reportsStart/startPageAccountant?pageUUID=...&userId=...&uuidValue=...

        The important outcome: a fresh pageUUID/uuidValue pair suitable as seed for remainGoodsReport/startPage.
        """
        user_id = (base_ctx.get("userId") or ITIGRIS_WEB_USER_ID or "").strip()
        page_uuid = (base_ctx.get("pageUUID") or "").strip()
        uuid_value = (base_ctx.get("uuidValue") or "").strip()
        debug: Dict[str, Any] = {
            "input": {"userId": user_id or None, "pageUUID": page_uuid or None, "uuidValue": uuid_value or None},
            "steps": [],
        }
        if not (user_id and page_uuid and uuid_value):
            debug["error"] = "missing_base_ctx"
            return {"_debug": debug}  # type: ignore[return-value]

        def _ctx_from_url(u: str) -> Dict[str, str]:
            try:
                absolute = urljoin(f"https://optima.itigris.ru/{APP_NAME}/", u.lstrip("/"))
                parsed = urlparse(absolute)
                qs = parse_qs(parsed.query)
                return {
                    "userId": (qs.get("userId") or [""])[0],
                    "pageUUID": (qs.get("pageUUID") or [""])[0],
                    "uuidValue": (qs.get("uuidValue") or [""])[0],
                    "companyUUID": (qs.get("companyUUID") or [company_uuid])[0] or company_uuid,
                }
            except Exception:
                return {}

        # 1) openMenuElement (POST) -> 302
        open1_url = f"https://optima.itigris.ru/{APP_NAME}/menu/openMenuElement?menuItem=accountantReports"
        payload = {"userId": user_id, "pageUUID": page_uuid, "uuidValue": uuid_value, "companyUUID": company_uuid}
        r1 = await client.post(open1_url, data=payload, headers=web_headers, follow_redirects=False)
        loc1 = r1.headers.get("location") or r1.headers.get("Location") or ""
        ctx1 = _ctx_from_url(loc1) if loc1 else {}
        debug["steps"].append(
            {
                "step": "openMenuElement",
                "status": r1.status_code,
                "location": loc1[:500] if loc1 else None,
                "ctx_from_location": {k: (ctx1.get(k) or None) for k in ["userId", "pageUUID", "uuidValue", "companyUUID"]},
            }
        )
        if r1.status_code not in {302, 303} or not loc1:
            debug["error"] = "openMenuElement_no_redirect"
            return {"_debug": debug}  # type: ignore[return-value]

        # Sometimes the new page context is already present in redirect URL.
        # 2) openMenuElement2 (GET) -> 302
        r2 = await _follow_location(client, loc1)
        loc2 = r2.headers.get("location") or r2.headers.get("Location") or ""
        ctx2 = _ctx_from_url(loc2) if loc2 else {}
        debug["steps"].append(
            {
                "step": "openMenuElement2",
                "status": r2.status_code,
                "location": loc2[:500] if loc2 else None,
                "ctx_from_location": {k: (ctx2.get(k) or None) for k in ["userId", "pageUUID", "uuidValue", "companyUUID"]},
            }
        )
        if r2.status_code not in {302, 303} or not loc2:
            debug["error"] = "openMenuElement2_no_redirect"
            return {"_debug": debug}  # type: ignore[return-value]

        # 3) startPageAccountant (GET) -> 200 HTML
        r3 = await _follow_location(client, loc2)
        debug["steps"].append(
            {
                "step": "startPageAccountant",
                "status": r3.status_code,
                "final_url": str(r3.request.url)[:500] if getattr(r3, "request", None) else None,
            }
        )
        if r3.status_code != 200:
            debug["error"] = "startPageAccountant_not_200"
            return {"_debug": debug}  # type: ignore[return-value]

        # Prefer context from the final URL query (closest to what browser uses), fallback to HTML.
        final_url = str(r3.request.url)
        parsed = urlparse(final_url)
        qs = parse_qs(parsed.query)
        out = {
            "userId": (qs.get("userId") or [ctx2.get("userId") or ctx1.get("userId") or user_id])[0] or user_id,
            "pageUUID": (qs.get("pageUUID") or [ctx2.get("pageUUID") or ctx1.get("pageUUID") or page_uuid])[0] or page_uuid,
            "uuidValue": (qs.get("uuidValue") or [ctx2.get("uuidValue") or ctx1.get("uuidValue") or uuid_value])[0] or uuid_value,
            "companyUUID": company_uuid,
        }
        html_ctx = _extract_optima_ctx_from_text(r3.text or "")
        debug["steps"].append(
            {
                "step": "startPageAccountant_extract",
                "ctx_from_final_url": {k: (out.get(k) or None) for k in ["userId", "pageUUID", "uuidValue", "companyUUID"]},
                "ctx_from_html": {k: (html_ctx.get(k) or None) for k in ["userId", "pageUUID", "uuidValue", "companyUUID"]},
            }
        )
        if html_ctx.get("pageUUID") and html_ctx.get("uuidValue"):
            out["pageUUID"] = html_ctx.get("pageUUID") or out["pageUUID"]
            out["uuidValue"] = html_ctx.get("uuidValue") or out["uuidValue"]
        out["_debug"] = debug  # type: ignore[typeddict-item]
        return out

    async with httpx.AsyncClient(timeout=TIMEOUT, follow_redirects=True) as client:
        # Preferred: server-side login to get fresh session.
        if ITIGRIS_WEB_LOGIN and ITIGRIS_WEB_PASSWORD and ITIGRIS_WEB_KEY:
            ctx = await login_and_get_context(client)
            # remainGoodsReport uses its own pageUUID/uuidValue chain:
            # login -> userStart -> remainGoodsReport/startPage -> reportPage (prepareData=true) -> reportPage.
            report_user_id = (ITIGRIS_WEB_USER_ID or ctx.get("userId") or REMAINGOODS_WEB_USER_ID or "").strip()
            if not report_user_id:
                raise HTTPException(status_code=502, detail={"error": "auto_web_missing_user_id_for_reports"})

            # 0) Initialize navigation context for reports as seen in HAR.
            # This yields the correct seed pageUUID/uuidValue expected by remainGoodsReport/startPage.
            module_ctx: Dict[str, str] = {}
            try:
                module_ctx = await _init_accountant_reports_ctx(client, ctx)
            except Exception:
                module_ctx = {}

            report_seed_page_uuid = (
                (module_ctx.get("pageUUID") or "").strip()
                or (ITIGRIS_REMAINGOODSREPORT_PAGE_UUID or "").strip()
                or (ITIGRIS_WEB_PAGE_UUID or "").strip()
                or (ctx.get("pageUUID") or "").strip()
                or ""
            )
            report_seed_uuid_value = (
                (module_ctx.get("uuidValue") or "").strip()
                or (ITIGRIS_REMAINGOODSREPORT_UUID_VALUE or "").strip()
                or (ITIGRIS_WEB_UUID_VALUE or "").strip()
                or (ctx.get("uuidValue") or "").strip()
                or ""
            )

            # 1) startPage (initializes report pageUUID/uuidValue)
            start_url = f"https://optima.itigris.ru/{APP_NAME}/remainGoodsReport/startPage"
            async def _post_start_page(seed_page_uuid: str, seed_uuid_value: str) -> httpx.Response:
                payload = {
                    "userId": report_user_id,
                    "uuidValue": seed_uuid_value,
                    "pageUUID": seed_page_uuid,
                    "companyUUID": company_uuid,
                }
                return await client.post(start_url, data=payload, headers=web_headers)

            # startPage expects the reports navigation context (HAR chain).
            preferred_seed_page_uuid = (
                report_seed_page_uuid or ""
            )
            preferred_seed_uuid_value = (
                report_seed_uuid_value or ""
            )

            start_payload_used = {"pageUUID": preferred_seed_page_uuid, "uuidValue": preferred_seed_uuid_value}
            start_resp = await _post_start_page(start_payload_used["pageUUID"], start_payload_used["uuidValue"])

            # startPage may legitimately return 211 ("Повторный запрос") while the page/session context settles.
            # Retry with backoff; also attempt to refresh the reports navigation context.
            for attempt in range(1, 7):
                if start_resp.status_code == 200:
                    break
                if start_resp.status_code == 211:
                    try:
                        refreshed = await _init_accountant_reports_ctx(client, ctx)
                        if refreshed.get("pageUUID") and refreshed.get("uuidValue"):
                            module_ctx = refreshed
                            start_payload_used["pageUUID"] = refreshed["pageUUID"]
                            start_payload_used["uuidValue"] = refreshed["uuidValue"]
                    except Exception:
                        pass
                    await asyncio.sleep(min(2.0, 0.2 * (2 ** (attempt - 1))))
                    start_resp = await _post_start_page(start_payload_used["pageUUID"], start_payload_used["uuidValue"])
                    continue
                break

            if start_resp.status_code != 200:
                ctx["_report_ctx"] = {  # type: ignore[typeddict-item]
                    "report_user_id": report_user_id,
                    "start_status": start_resp.status_code,
                    "pageUUID_seed": start_payload_used["pageUUID"],
                    "uuidValue_seed": start_payload_used["uuidValue"],
                    "module_ctx": module_ctx,
                    "login_debug": ctx.get("_debug"),
                    "module_overrides_present": bool(
                        ITIGRIS_REMAINGOODSREPORT_PAGE_UUID or ITIGRIS_REMAINGOODSREPORT_UUID_VALUE
                    ),
                }
                raise HTTPException(
                    status_code=502,
                    detail={
                        "error": "auto_web_startPage_failed",
                        "status_code": start_resp.status_code,
                        "body_snippet": (start_resp.text or "")[:1000],
                        "report_debug": ctx.get("_report_ctx"),
                    },
                )
            start_ctx = _extract_optima_ctx_from_text(start_resp.text or "")
            report_page_uuid = (start_ctx.get("pageUUID") or "").strip() or start_payload_used["pageUUID"] or ""
            report_uuid_value = (start_ctx.get("uuidValue") or "").strip() or start_payload_used["uuidValue"] or ""

            # 2) reportPage prepareData=true (usually updates uuidValue)
            prep_resp: httpx.Response
            prep_ctx: Dict[str, str] = {}

            # Prefer labeled values from startPage response. If extractor can't find them, we'll time out
            # and fall back to UUID candidates from startPage (validated by the prepare response).
            labeled_page_uuid = (start_ctx.get("pageUUID") or "").strip()
            labeled_uuid_value = (start_ctx.get("uuidValue") or "").strip()
            if labeled_page_uuid:
                report_page_uuid = labeled_page_uuid
            if labeled_uuid_value:
                report_uuid_value = labeled_uuid_value

            prep_debug: Dict[str, Any] = {
                "pageUUID": report_page_uuid,
                "uuidValue": report_uuid_value,
                "attempts": [],
            }

            prepare_ok = False
            for attempt in range(1, 13):
                prep_form = build_report_form(user_id=report_user_id, page_uuid=report_page_uuid, uuid_value=report_uuid_value)
                prep_form["prepareData"] = "true"
                prep_resp = await client.post(url, data=prep_form, headers=web_headers)
                if prep_resp.status_code == 200:
                    prep_ctx = _extract_optima_ctx_from_text(prep_resp.text or "")
                    # pageUUID should remain stable; uuidValue may update.
                    report_uuid_value = (prep_ctx.get("uuidValue") or "").strip() or report_uuid_value
                    prepare_ok = True
                    break
                if prep_resp.status_code == 211:
                    sleep_s = min(3.0, 0.25 * (2 ** (attempt - 1)))
                    prep_debug["attempts"].append(
                        {
                            "attempt": attempt,
                            "status_code": 211,
                            "sleep_s": sleep_s,
                            "payload_meta": {
                                "userId": prep_form.get("userId"),
                                "pageUUID": prep_form.get("pageUUID"),
                                "uuidValue": prep_form.get("uuidValue"),
                                "prepareData": prep_form.get("prepareData"),
                                "groupByDepartment": prep_form.get("groupByDepartment"),
                                "departments_count": len(dep_ids),
                            },
                        }
                    )
                    await asyncio.sleep(sleep_s)
                    continue

                cookie_names = sorted({c.name for c in client.cookies.jar})
                ctx["_report_ctx"] = {  # type: ignore[typeddict-item]
                    "report_user_id": report_user_id,
                    "start_status": start_resp.status_code,
                    "pageUUID": report_page_uuid,
                    "uuidValue": report_uuid_value,
                    "seed_pageUUID": report_seed_page_uuid,
                    "seed_uuidValue": report_seed_uuid_value,
                    "start_ctx": start_ctx,
                    "start_snippet": (start_resp.text or "")[:400],
                    "module_ctx": module_ctx,
                    "login_debug": ctx.get("_debug"),
                    "cookie_names": cookie_names,
                    "prepare_debug": prep_debug,
                }
                raise HTTPException(
                    status_code=502,
                    detail={
                        "error": "auto_web_prepare_failed",
                        "status_code": prep_resp.status_code,
                        "content_type": (prep_resp.headers.get("content-type") or ""),
                        "body_snippet": (prep_resp.text or "")[:1200],
                        "report_debug": ctx.get("_report_ctx"),
                    },
                )

            if not prepare_ok:
                ctx["_report_ctx"] = {  # type: ignore[typeddict-item]
                    "report_user_id": report_user_id,
                    "start_status": start_resp.status_code,
                    "pageUUID": report_page_uuid,
                    "uuidValue": report_uuid_value,
                    "seed_pageUUID": report_seed_page_uuid,
                    "seed_uuidValue": report_seed_uuid_value,
                    "start_ctx": start_ctx,
                    "start_snippet": (start_resp.text or "")[:400],
                    "prepare_debug": prep_debug,
                    "module_ctx": module_ctx,
                }
                raise HTTPException(
                    status_code=502,
                    detail={
                        "error": "auto_web_prepare_timeout",
                        "status_code": prep_resp.status_code,
                        "content_type": (prep_resp.headers.get("content-type") or ""),
                        "body_snippet": (prep_resp.text or "")[:1200],
                        "report_debug": ctx.get("_report_ctx"),
                    },
                )

            # 3) reportPage final (parse this HTML)
            final_resp: httpx.Response
            for attempt in range(1, 9):
                final_resp = await do_report_request(client, report_user_id, report_page_uuid, report_uuid_value)
                if final_resp.status_code == 200:
                    break
                if final_resp.status_code == 211:
                    # In practice Optima sometimes expects a *fresh* uuidValue for the final request
                    # even though prepareData=true succeeded (browser often sends a new one).
                    # We don't always have a reliable extractor for it, so we generate a new UUID and retry.
                    report_uuid_value = str(uuid.uuid4())
                    await asyncio.sleep(min(2.0, 0.2 * (2 ** (attempt - 1))))
                    continue
                break
            # expose these values for debug below
            ctx["_report_ctx"] = {  # type: ignore[typeddict-item]
                "report_user_id": report_user_id,
                "start_status": start_resp.status_code,
                "prep_status": prep_resp.status_code,
                "final_status": final_resp.status_code,
                "pageUUID": report_page_uuid,
                "uuidValue": report_uuid_value,
                "module_ctx": module_ctx,
                "seed_pageUUID": report_seed_page_uuid,
                "seed_uuidValue": report_seed_uuid_value,
                "start_ctx": start_ctx,
                "start_snippet": (start_resp.text or "")[:400],
                "prep_ctx": prep_ctx,
                "prep_snippet": (prep_resp.text or "")[:400],
                "final_snippet": (final_resp.text or "")[:400],
            }
            resp = final_resp
        else:
            # Fallback: static Cookie header + static context from env.
            if not REMAINGOODS_WEB_COOKIE:
                raise HTTPException(status_code=500, detail="auto_web_cookie_not_configured")
            if not REMAINGOODS_WEB_USER_ID or not REMAINGOODS_WEB_PAGE_UUID or not REMAINGOODS_WEB_UUID_VALUE:
                raise HTTPException(status_code=500, detail="auto_web_payload_not_configured")
            client.headers.update({"Cookie": REMAINGOODS_WEB_COOKIE})
            resp = await do_report_request(client, REMAINGOODS_WEB_USER_ID, REMAINGOODS_WEB_PAGE_UUID, REMAINGOODS_WEB_UUID_VALUE)

    content_type = (resp.headers.get("content-type") or "").lower()
    text_snippet = resp.text[:2000] if resp.text else ""
    looks_like_login = any(
        marker in (resp.text or "").lower()
        for marker in [
            "jsessionid",
            "login",
            "авторизац",
            "войти",
            "password",
            "username",
            "j_username",
            "j_password",
        ]
    ) and ("remainGoodsReport" not in (resp.text or ""))

    if resp.status_code != 200:
        raise HTTPException(
            status_code=502,
            detail={
                "error": f"auto_web_http_{resp.status_code}",
                "status_code": resp.status_code,
                "content_type": content_type,
                "looks_like_login_page": looks_like_login,
                "html_snippet": text_snippet,
                "cookies_present": sorted({c.name for c in client.cookies.jar}) if "client" in locals() else [],
                "report_context": {
                    "userId_used": (ctx.get("_report_ctx") or {}).get("report_user_id") if "ctx" in locals() else None,
                    "pageUUID_used": (ctx.get("_report_ctx") or {}).get("pageUUID") if "ctx" in locals() else None,
                    "uuidValue_used": (ctx.get("_report_ctx") or {}).get("uuidValue") if "ctx" in locals() else None,
                    "login_debug": ctx.get("_debug") if "ctx" in locals() else None,
                    "report_debug": ctx.get("_report_ctx") if "ctx" in locals() else None,
                }
            },
        )

    try:
        rows = _parse_remain_goods_html(resp.content)
    except HTTPException as e:
        raise HTTPException(
            status_code=502,
            detail={
                "error": "auto_web_parse_failed",
                "status_code": resp.status_code,
                "content_type": content_type,
                "looks_like_login_page": looks_like_login,
                "html_snippet": text_snippet,
                "parser_detail": e.detail,
            },
        )
    return f"remainGoodsReport_reportPage_{date_ddmmyyyy}.html", rows


def _parse_remain_goods_report_auto(filename: Optional[str], raw: bytes) -> Tuple[str, List[Dict[str, Any]]]:
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        return filename or "remainGoodsReport.csv", parse_remain_goods_csv(raw)
    if fn.endswith(".xlsx"):
        return filename or "remainGoodsReport.xlsx", parse_remain_goods_xlsx(raw)
    if fn.endswith(".xls"):
        return filename or "remainGoodsReport.xls", parse_remain_goods_xls(raw)

    # Heuristic fallback: try XLS, then XLSX, then CSV.
    try:
        return (filename or "remainGoodsReport.xls"), parse_remain_goods_xls(raw)
    except Exception:
        pass
    try:
        return (filename or "remainGoodsReport.xlsx"), parse_remain_goods_xlsx(raw)
    except Exception:
        pass
    return (filename or "remainGoodsReport.csv"), parse_remain_goods_csv(raw)


async def maybe_refresh_global_snapshot_from_itigris(force: bool = False) -> Dict[str, Any]:
    """
    Best-effort refresh of global snapshot from ITigris.
    Methods:
    1) direct file URL (ITIGRIS_REMAINGOODSREPORT_URL_TEMPLATE) [preferred]
    2) web session + reportPage HTML (ITIGRIS_REMAINGOODSREPORT_WEB_*) [fallback]
    """
    now = int(time.time())
    _contactlenses_auto_fetch_state["last_attempt_unix"] = now

    login_mode_configured = bool(ITIGRIS_WEB_LOGIN and ITIGRIS_WEB_PASSWORD and ITIGRIS_WEB_KEY)
    if not REMAINGOODS_AUTO_FETCH_URL_TEMPLATE and not REMAINGOODS_WEB_COOKIE and not login_mode_configured:
        _contactlenses_auto_fetch_state["last_error"] = "auto_fetch_not_configured"
        _contactlenses_auto_fetch_state["last_error_at_unix"] = now
        return {"ok": False, "error": "auto_fetch_not_configured"}

    global_snap = get_global_snapshot()
    if not force and global_snap:
        age = now - int(global_snap["stored_at_unix"])
        if age < REMAINGOODS_AUTO_REFRESH_MIN_SECONDS:
            return {"ok": True, "skipped": True, "reason": "recent_snapshot"}

    resolved_filename: str
    rows: List[Dict[str, Any]]
    method_used: str
    method_attempted: str = "url_template" if REMAINGOODS_AUTO_FETCH_URL_TEMPLATE else "web_reportPage"
    try:
        if REMAINGOODS_AUTO_FETCH_URL_TEMPLATE:
            filename, raw, err = await _auto_fetch_remain_goods_report_bytes()
            if err or not raw:
                raise HTTPException(status_code=502, detail=err or "auto_fetch_failed")
            resolved_filename, rows = _parse_remain_goods_report_auto(filename, raw)
            method_used = "url_template"
        else:
            resolved_filename, rows = await _auto_fetch_remain_goods_report_via_web()
            method_used = "web_reportPage"

        overall = remain_goods_totals(rows, prefer_summary_row=True)
        by_dep = remain_goods_totals_by_department(rows)
    except HTTPException as e:
        # Keep the detail for debugging (can be dict).
        _contactlenses_auto_fetch_state["last_error"] = e.detail
        _contactlenses_auto_fetch_state["last_error_at_unix"] = now
        return {
            "ok": False,
            "error": "auto_refresh_failed",
            "method_attempted": method_attempted,
            "impl_version": AUTO_REFRESH_IMPL_VERSION,
            "build_commit": BUILD_COMMIT or None,
            "debug": e.detail,
        }
    except Exception as exc:
        _contactlenses_auto_fetch_state["last_error"] = {"type": type(exc).__name__, "message": str(exc)}
        _contactlenses_auto_fetch_state["last_error_at_unix"] = now
        return {
            "ok": False,
            "error": "parse_error_unknown",
            "method_attempted": method_attempted,
            "impl_version": AUTO_REFRESH_IMPL_VERSION,
            "build_commit": BUILD_COMMIT or None,
            "debug": {"type": type(exc).__name__, "message": str(exc)},
        }

    expires = now + REMAINGOODS_SNAPSHOT_TTL_SECONDS
    global _contactlenses_report_global_snapshot
    _contactlenses_report_global_snapshot = {
        "stored_at_unix": now,
        "expires_at_unix": expires,
        "filename": resolved_filename,
        "rows_count": len(rows),
        "overall_summary": overall,
        "by_department": by_dep,
    }

    _contactlenses_auto_fetch_state["last_success_unix"] = now
    _contactlenses_auto_fetch_state["last_error"] = None
    _contactlenses_auto_fetch_state["last_error_at_unix"] = None
    _contactlenses_auto_fetch_state["last_filename"] = resolved_filename

    return {
        "ok": True,
        "skipped": False,
        "method": method_used,
        "impl_version": AUTO_REFRESH_IMPL_VERSION,
        "build_commit": BUILD_COMMIT or None,
        "stored_at_unix": now,
        "expires_at_unix": expires,
        "filename": resolved_filename,
        "rows_count": len(rows),
        "departments_count": len(by_dep),
        "overall_summary": overall,
    }


def remain_goods_totals_by_department(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        dep = str(r.get("department") or "").strip()
        if not dep:
            continue
        if normalize_header(dep) in {"итого", "всего", "итог", "total"}:
            continue
        grouped.setdefault(dep, []).append(r)

    result: Dict[str, Dict[str, Any]] = {}
    for dep, dep_rows in grouped.items():
        result[dep] = remain_goods_totals(dep_rows, prefer_summary_row=False)
    return result

def rows_to_excel_bytes(rows: List[Dict[str, Any]], sheet_name: str = "Remains") -> bytes:
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet(sheet_name[:31] or "Remains")
    data = rows or [{"message": "Нет данных"}]
    headers = list(data[0].keys())
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)
    for row_index, row in enumerate(data, start=1):
        for col, header in enumerate(headers):
            worksheet.write(row_index, col, row.get(header))
    workbook.close()
    output.seek(0)
    return output.getvalue()


async def fetch_optima_remains_once(
    category: str,
    department_id: Optional[int],
    page: int,
    filter_payload: Optional[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    if not REMOTE_API_KEY:
        raise RuntimeError("ITIGRIS_REMOTE_API_KEY or ITIGRIS_API_KEY is not configured")

    body: Dict[str, Any] = {"product": category, "page": page}
    if department_id:
        body["departmentId"] = department_id
    if filter_payload:
        body["filter"] = filter_payload

    async with httpx.AsyncClient(timeout=TIMEOUT) as client:
        response = await client.post(REMOTE_REMAINS_URL, params={"key": REMOTE_API_KEY}, json=body)
        if response.status_code in {404, 405} and not filter_payload:
            params = {"key": REMOTE_API_KEY, "product": category, "page": page}
            if department_id:
                params["departmentId"] = department_id
            response = await client.get(REMOTE_REMAINS_URL, params=params)

    if response.status_code != 200:
        raise RuntimeError(f"ITigris remoteRemains error {response.status_code}: {response.text[:500]}")
    data = response.json()
    if not isinstance(data, list):
        raise RuntimeError(f"Unexpected ITigris response: {data}")
    return data


async def fetch_optima_remains(
    category: str,
    department_ids: List[int],
    filter_payload: Optional[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    all_rows: List[Dict[str, Any]] = []
    for dep_id in department_ids:
        for page in range(1, 501):
            rows = await fetch_optima_remains_once(category, dep_id, page, filter_payload)
            if not rows:
                break
            for row in rows:
                row.setdefault("departmentId", dep_id)
            all_rows.extend(rows)
        else:
            raise RuntimeError(f"Pagination limit reached for department {dep_id}")
    return all_rows


def summarize_rows(category: str, rows: List[Dict[str, Any]], limit: int = 0) -> Dict[str, Any]:
    if category == "contactlenses":
        totals = contact_lenses_totals(rows)
        total_qty = int(totals["total_qty_packs"])
        total_value = float(totals["total_value"])
        avg_price = total_value / total_qty if total_qty else 0.0
    else:
        total_qty, total_value = sum_qty_value(rows)
        avg_price = total_value / total_qty if total_qty else 0.0

    response: Dict[str, Any] = {
        "category": category,
        "source": "ITigris remoteRemains/list",
        "positions_count": len(rows),
        "summary": {
            "total_qty": total_qty,
            "total_value": round(total_value, 2),
            "avg_price": round(avg_price, 2),
        },
    }
    if category == "contactlenses":
        response["summary"].update(
            {
                "total_qty_packs": int(totals["total_qty_packs"]),
                "total_qty_units": int(totals["total_qty_units"]),
                "open_packs_detected": int(totals["open_packs_detected"]),
                "note": (
                    "contactlenses totals are reported as packs+units when possible. "
                    "remoteRemains may undercount open packs compared to report exports."
                ),
            }
        )
    if limit > 0:
        response["items"] = rows[:limit]
        response["items_truncated"] = len(rows) > limit
        response["items_total"] = len(rows)
    return response


def build_breakdown(category: str, rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, int]]:
    result: Dict[str, Dict[str, int]] = {}
    for field in BREAKDOWN_FIELDS.get(category, CATEGORY_FILTERS[category]):
        bucket: Dict[str, int] = {}
        for row in rows:
            value = row.get(field)
            if value in (None, ""):
                continue
            bucket[str(value)] = bucket.get(str(value), 0) + amount_as_int(row)
        result[field] = dict(sorted(bucket.items(), key=lambda item: item[1], reverse=True))
    return result


@app.get("/healthz")
def healthz() -> Dict[str, Any]:
    return {
        "ok": True,
        "version": app.version,
        "build_commit": BUILD_COMMIT or None,
        "auto_refresh_impl_version": AUTO_REFRESH_IMPL_VERSION,
        "itigris_app": APP_NAME,
        "remote_api_key_configured": bool(REMOTE_API_KEY),
        "external_api_key_configured": bool(EXTERNAL_API_KEY),
        "server_token_configured": bool(ODL_SERVER_TOKEN),
    }


@app.get("/", include_in_schema=False)
def home(request: Request) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err
    return {"message": "ODL Sales Analyst MVP API is running", "version": app.version}


@app.get("/departments")
def departments(request: Request) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err
    return {
        "departments": [{"name": name, "id": dep_id} for name, dep_id in DEPARTMENTS.items()],
        "aliases": DEPARTMENT_ALIASES,
        "groups": GROUPS_RU,
    }


@app.get("/categories")
def categories(request: Request) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err
    return {
        "codes": list(CATEGORY_FILTERS.keys()),
        "aliases": CATEGORY_ALIASES,
        "filters": CATEGORY_FILTERS,
        "filter_aliases": FILTER_FIELD_ALIASES,
    }


@app.post("/remains-filtered")
async def remains_filtered(request: Request, body: RemainsFilteredRequest) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err

    category = normalize_category(body.category)
    if category not in CATEGORY_FILTERS:
        return JSONResponse({"error": "unknown_category", "category": body.category}, status_code=400)

    try:
        dep_ids = resolve_dep_ids(body.department_id, body.department_name, body.group)
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)

    normalized_filters = normalize_filters(category, body.filters)
    filter_payload = build_filter_payload(normalized_filters, body.min_price, body.max_price, body.price)
    try:
        rows = await fetch_optima_remains(category, dep_ids, filter_payload)
    except Exception as exc:
        return JSONResponse({"error": "upstream_error", "detail": str(exc)}, status_code=502)

    limit = body.items_limit if body.return_items else 0
    response = summarize_rows(category, rows, limit=limit)
    response["scope"] = {
        "department_id": body.department_id,
        "department_name": body.department_name,
        "group": body.group,
        "departments_used": dep_ids,
    }
    response["filters_used"] = filter_payload or {}
    if normalized_filters and "_ignoredFilters" in normalized_filters:
        response["ignored_filters"] = normalized_filters["_ignoredFilters"]
    return response


@app.post("/count-by-filters")
async def count_by_filters(request: Request, body: RemainsFilteredRequest) -> Any:
    body.return_items = False
    body.items_limit = 0
    return await remains_filtered(request, body)


@app.get("/count-by-price/{category}")
async def count_by_price(
    request: Request,
    category: str,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    department_id: Optional[int] = None,
    department_name: Optional[str] = None,
    group: Optional[str] = None,
    limit_items: int = Query(50, ge=0, le=500),
) -> Any:
    body = RemainsFilteredRequest(
        category=category,
        department_id=department_id,
        department_name=department_name,
        group=group,
        min_price=min_price,
        max_price=max_price,
        return_items=limit_items > 0,
        items_limit=limit_items,
    )
    return await remains_filtered(request, body)


@app.get("/count/{category}")
async def count_category(
    request: Request,
    category: str,
    department_id: Optional[int] = None,
    department_name: Optional[str] = None,
    group: Optional[str] = None,
) -> Any:
    return await count_by_price(
        request=request,
        category=category,
        department_id=department_id,
        department_name=department_name,
        group=group,
        limit_items=0,
    )


@app.get("/breakdown/{category}")
async def breakdown_category(
    request: Request,
    category: str,
    department_id: Optional[int] = None,
    department_name: Optional[str] = None,
    group: Optional[str] = None,
) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err

    normalized_category = normalize_category(category)
    if normalized_category not in CATEGORY_FILTERS:
        return JSONResponse({"error": "unknown_category", "category": category}, status_code=400)
    try:
        dep_ids = resolve_dep_ids(department_id, department_name, group)
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)

    try:
        rows = await fetch_optima_remains(normalized_category, dep_ids, filter_payload=None)
    except Exception as exc:
        return JSONResponse({"error": "upstream_error", "detail": str(exc)}, status_code=502)

    total_qty, total_value = sum_qty_value(rows)
    return {
        "category": normalized_category,
        "scope": {
            "department_id": department_id,
            "department_name": department_name,
            "group": group,
            "departments_used": dep_ids,
        },
        "source": "ITigris remoteRemains/list",
        "total_qty": total_qty,
        "total_value": round(total_value, 2),
        "breakdown": build_breakdown(normalized_category, rows),
    }


@app.get("/gpt/breakdown/{category}")
async def gpt_breakdown(
    request: Request,
    category: str,
    department_id: Optional[int] = None,
    department_name: Optional[str] = None,
    group: Optional[str] = None,
    top_n: int = Query(10, ge=1, le=30),
) -> Any:
    result = await breakdown_category(request, category, department_id, department_name, group)
    if isinstance(result, JSONResponse):
        return result
    compact = {}
    for field, values in result["breakdown"].items():
        compact[field] = dict(list(values.items())[:top_n])
    result["breakdown"] = compact
    result["note"] = "Это остатки ITigris remoteRemains, не продажи."
    return result


@app.get("/remains/{category}", include_in_schema=False)
async def remains_excel(
    request: Request,
    category: str,
    department_id: Optional[int] = None,
    department_name: Optional[str] = None,
    group: Optional[str] = None,
) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err
    normalized_category = normalize_category(category)
    if normalized_category not in CATEGORY_FILTERS:
        return JSONResponse({"error": "unknown_category", "category": category}, status_code=400)
    try:
        dep_ids = resolve_dep_ids(department_id, department_name, group)
        rows = await fetch_optima_remains(normalized_category, dep_ids, filter_payload=None)
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"error": "upstream_error", "detail": str(exc)}, status_code=502)

    data = rows_to_excel_bytes(rows, sheet_name=normalized_category)
    filename = f"{normalized_category}_remains.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@app.post("/sales/analyze")
async def sales_analyze(request: Request, body: SalesAnalyzeRequest) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err
    dataset = body.data if body.data is not None else load_input_folder(Path("data/input"))
    return analyze_dataset(dataset)


@app.post("/sales/analyze-upload")
async def sales_analyze_upload(
    request: Request,
    sales_period: UploadFile = File(...),
    plan_fact: Optional[UploadFile] = File(None),
    employees: Optional[UploadFile] = File(None),
    categories: Optional[UploadFile] = File(None),
    training: Optional[UploadFile] = File(None),
) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err

    dataset = {
        "sales_period": await load_csv_upload(sales_period, required=True),
        "plan_fact": await load_csv_upload(plan_fact, required=False),
        "employees": await load_csv_upload(employees, required=False),
        "categories": await load_csv_upload(categories, required=False),
        "training": await load_csv_upload(training, required=False),
    }
    return analyze_dataset(dataset)


@app.post("/sales/analyze-csv")
async def sales_analyze_csv(request: Request, body: SalesAnalyzeCsvRequest) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err

    dataset = {
        "sales_period": csv_text_to_rows(body.sales_period_csv),
        "plan_fact": csv_text_to_rows(body.plan_fact_csv) if body.plan_fact_csv else [],
        "employees": csv_text_to_rows(body.employees_csv) if body.employees_csv else [],
        "categories": csv_text_to_rows(body.categories_csv) if body.categories_csv else [],
        "training": csv_text_to_rows(body.training_csv) if body.training_csv else [],
    }
    return analyze_dataset(dataset)


@app.post("/contactlenses/remainGoodsReport/analyze")
async def contactlenses_remain_goods_report_analyze(
    request: Request,
    report_file: UploadFile = File(...),
) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err

    raw = await report_file.read()
    if len(raw) > 30_000_000:
        raise HTTPException(status_code=413, detail="report_file_too_large")

    filename = (report_file.filename or "").lower()
    content_type = (report_file.content_type or "").lower()
    if filename.endswith(".csv") or "csv" in content_type or content_type == "text/plain":
        rows = parse_remain_goods_csv(raw)
    elif filename.endswith(".xlsx") or "spreadsheetml" in content_type:
        rows = parse_remain_goods_xlsx(raw)
    elif filename.endswith(".xls") or content_type in {"application/vnd.ms-excel"}:
        rows = parse_remain_goods_xls(raw)
    else:
        raise HTTPException(status_code=400, detail="unsupported_report_file_type")

    totals = remain_goods_totals(rows)
    return {
        "category": "contactlenses",
        "source": "remainGoodsReport upload (truth source for contact lenses packs/units/value)",
        "rows_count": len(rows),
        "summary": totals,
        "note": (
            "Use this endpoint for exact contact lenses stock in packs/units/value including open packs. "
            "ITigris remoteRemains is an API snapshot and may undercount open packs."
        ),
    }


@app.post("/contactlenses/remainGoodsReport/analyze-csv")
async def contactlenses_remain_goods_report_analyze_csv(
    request: Request,
    body: RemainGoodsReportCsvRequest,
) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err
    raw = body.report_csv.encode("utf-8")
    if len(raw) > 30_000_000:
        raise HTTPException(status_code=413, detail="report_csv_too_large")
    rows = parse_remain_goods_csv(raw)
    totals = remain_goods_totals(rows)
    return {
        "category": "contactlenses",
        "source": "remainGoodsReport CSV text (truth source for contact lenses packs/units/value)",
        "rows_count": len(rows),
        "summary": totals,
        "note": (
            "Use this endpoint for exact contact lenses stock in packs/units/value including open packs. "
            "ITigris remoteRemains is an API snapshot and may undercount open packs."
        ),
    }


@app.post("/contactlenses/remainGoodsReport/analyze-base64")
async def contactlenses_remain_goods_report_analyze_base64(
    request: Request,
    body: RemainGoodsReportBase64Request,
) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err

    try:
        raw = base64.b64decode(body.file_base64, validate=True)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid_base64")
    if len(raw) > 30_000_000:
        raise HTTPException(status_code=413, detail="report_file_too_large")

    filename = (body.filename or "").lower()
    if filename.endswith(".csv"):
        rows = parse_remain_goods_csv(raw)
    elif filename.endswith(".xlsx"):
        rows = parse_remain_goods_xlsx(raw)
    elif filename.endswith(".xls"):
        rows = parse_remain_goods_xls(raw)
    else:
        raise HTTPException(status_code=400, detail="unsupported_report_file_type")

    totals = remain_goods_totals(rows)
    return {
        "category": "contactlenses",
        "source": "remainGoodsReport base64 (truth source for contact lenses packs/units/value)",
        "rows_count": len(rows),
        "summary": totals,
        "note": (
            "Use this endpoint for exact contact lenses stock in packs/units/value including open packs. "
            "ITigris remoteRemains is an API snapshot and may undercount open packs."
        ),
    }


@app.post("/contactlenses/remainGoodsReport/snapshot/set")
async def contactlenses_remain_goods_report_snapshot_set(
    request: Request,
    body: RemainGoodsReportSnapshotSetRequest,
) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err

    dep_id = normalize_department(None, body.department_name)
    if not dep_id:
        raise HTTPException(status_code=400, detail="unknown_department")

    try:
        raw = base64.b64decode(body.file_base64, validate=True)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid_base64")
    if len(raw) > 30_000_000:
        raise HTTPException(status_code=413, detail="report_file_too_large")

    rows = parse_remain_goods_by_filename(body.filename, raw)
    totals = remain_goods_totals(rows, prefer_summary_row=True)
    now = int(time.time())
    expires = now + REMAINGOODS_SNAPSHOT_TTL_SECONDS
    _contactlenses_report_snapshots[dep_id] = {
        "department_id": dep_id,
        "department_name": body.department_name,
        "stored_at_unix": now,
        "expires_at_unix": expires,
        "filename": body.filename,
        "rows_count": len(rows),
        "summary": totals,
    }
    return {
        "ok": True,
        "snapshot": RemainGoodsReportSnapshotInfo(
            department_id=dep_id,
            department_name=body.department_name,
            stored_at_unix=now,
            expires_at_unix=expires,
            filename=body.filename,
            rows_count=len(rows),
        ).model_dump(),
        "summary": totals,
    }


@app.post("/contactlenses/remainGoodsReport/snapshot/set-global")
async def contactlenses_remain_goods_report_snapshot_set_global(
    request: Request,
    body: RemainGoodsReportGlobalSnapshotSetRequest,
) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err

    try:
        raw = base64.b64decode(body.file_base64, validate=True)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid_base64")
    if len(raw) > 30_000_000:
        raise HTTPException(status_code=413, detail="report_file_too_large")

    rows = parse_remain_goods_by_filename(body.filename, raw)
    overall = remain_goods_totals(rows, prefer_summary_row=True)
    by_dep = remain_goods_totals_by_department(rows)

    now = int(time.time())
    expires = now + REMAINGOODS_SNAPSHOT_TTL_SECONDS
    global _contactlenses_report_global_snapshot
    _contactlenses_report_global_snapshot = {
        "stored_at_unix": now,
        "expires_at_unix": expires,
        "filename": body.filename,
        "rows_count": len(rows),
        "overall_summary": overall,
        "by_department": by_dep,
    }
    return {
        "ok": True,
        "snapshot": {
            "stored_at_unix": now,
            "expires_at_unix": expires,
            "filename": body.filename,
            "rows_count": len(rows),
            "departments_count": len(by_dep),
        },
        "overall_summary": overall,
        "departments": list(by_dep.keys()),
    }


@app.get("/contactlenses/remainGoodsReport/auto/status")
async def contactlenses_remain_goods_report_auto_status(request: Request) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err
    global_snap = get_global_snapshot()
    return {
        "build_commit": BUILD_COMMIT or None,
        "auto_refresh_impl_version": AUTO_REFRESH_IMPL_VERSION,
        "auto_fetch_configured": bool(REMAINGOODS_AUTO_FETCH_URL_TEMPLATE or REMAINGOODS_WEB_COOKIE or (ITIGRIS_WEB_LOGIN and ITIGRIS_WEB_PASSWORD and ITIGRIS_WEB_KEY)),
        "auto_refresh_min_seconds": REMAINGOODS_AUTO_REFRESH_MIN_SECONDS,
        "auto_fetch_methods": {
            "url_template": bool(REMAINGOODS_AUTO_FETCH_URL_TEMPLATE),
            "web_reportPage_cookie": bool(REMAINGOODS_WEB_COOKIE),
            "web_reportPage_login": bool(ITIGRIS_WEB_LOGIN and ITIGRIS_WEB_PASSWORD and ITIGRIS_WEB_KEY),
        },
        "remote_api_key_configured": bool(REMOTE_API_KEY),
        "external_api_key_configured": bool(EXTERNAL_API_KEY),
        "state": dict(_contactlenses_auto_fetch_state),
        "global_snapshot_present": bool(global_snap),
        "global_snapshot": (
            {
                "stored_at_unix": global_snap["stored_at_unix"],
                "expires_at_unix": global_snap["expires_at_unix"],
                "filename": global_snap["filename"],
                "rows_count": global_snap["rows_count"],
                "departments_count": len(global_snap["by_department"]),
            }
            if global_snap
            else None
        ),
    }


@app.post("/contactlenses/remainGoodsReport/auto/refresh")
async def contactlenses_remain_goods_report_auto_refresh(request: Request, force: bool = Query(False)) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err
    status = await maybe_refresh_global_snapshot_from_itigris(force=force)
    if not status.get("ok"):
        # Return debug payload to quickly fix cookie/login/export parsing issues.
        raise HTTPException(status_code=502, detail=status)
    return status


@app.get("/contactlenses/stock/{department_name}")
async def contactlenses_stock(
    request: Request,
    department_name: str,
    source: str = Query("auto", description="auto|snapshot|api"),
) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err

    dep_id = normalize_department(None, department_name)
    if not dep_id:
        return JSONResponse({"error": "unknown_department"}, status_code=400)

    if source in {"auto", "snapshot"}:
        # 0) Best-effort refresh from ITigris export URL (if configured).
        if source == "auto" and (
            REMAINGOODS_AUTO_FETCH_URL_TEMPLATE
            or REMAINGOODS_WEB_COOKIE
            or (ITIGRIS_WEB_LOGIN and ITIGRIS_WEB_PASSWORD and ITIGRIS_WEB_KEY)
        ):
            await maybe_refresh_global_snapshot_from_itigris(force=False)

        # 1) Prefer global snapshot (single upload for all departments).
        global_snap = get_global_snapshot()
        if global_snap:
            # Match by normalized department name.
            for dep_name, summary in global_snap["by_department"].items():
                if normalize_header(dep_name) == normalize_header(department_name):
                    return {
                        "category": "contactlenses",
                        "department": department_name,
                        "source": "remainGoodsReport global snapshot (truth for packs/units/value; includes open packs)",
                        "snapshot": {
                            "stored_at_unix": global_snap["stored_at_unix"],
                            "expires_at_unix": global_snap["expires_at_unix"],
                            "filename": global_snap["filename"],
                            "rows_count": global_snap["rows_count"],
                        },
                        "summary": summary,
                        "overall_summary": global_snap["overall_summary"],
                    }

        # 2) Per-department snapshot (older mode).
        snap = get_snapshot(dep_id)
        if snap:
            return {
                "category": "contactlenses",
                "department": department_name,
                "source": "remainGoodsReport snapshot (truth for packs/units/value; includes open packs)",
                "snapshot": {
                    "department_id": snap["department_id"],
                    "department_name": snap["department_name"],
                    "stored_at_unix": snap["stored_at_unix"],
                    "expires_at_unix": snap["expires_at_unix"],
                    "filename": snap["filename"],
                    "rows_count": snap["rows_count"],
                },
                "summary": snap["summary"],
            }
        if source == "snapshot":
            return JSONResponse(
                {
                    "error": "snapshot_missing",
                    "detail": "No remainGoodsReport snapshot stored for this department. Upload it via /contactlenses/remainGoodsReport/snapshot/set-global or /contactlenses/remainGoodsReport/snapshot/set.",
                },
                status_code=404,
            )

    if source in {"auto", "api"}:
        body = RemainsFilteredRequest(category="contactlenses", department_name=department_name, return_items=False, items_limit=0)
        api_result = await remains_filtered(request, body)
        if isinstance(api_result, JSONResponse):
            return api_result
        api_result["warning"] = (
            "remoteRemains is an API snapshot. For contact lenses it may undercount open packs and does not provide exact units. "
            "For exact packs+units+sum use remainGoodsReport snapshot."
        )
        return api_result

    return JSONResponse({"error": "bad_source", "allowed": ["auto", "snapshot", "api"]}, status_code=400)


@app.get("/openapi.json", include_in_schema=False)
def openapi_json(request: Request) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err

    schema = get_openapi(
        title=app.title,
        version=app.version,
        routes=app.routes,
        description=(
            "MVP API: ITigris remoteRemains for stock/assortment plus sales analytical report "
            "from DataLens/Google Sheets CSV input. remoteRemains is not sales."
        ),
    )
    schema["servers"] = [{"url": SERVER_URL}]
    schema.setdefault("components", {}).setdefault("securitySchemes", {})
    schema["components"]["securitySchemes"]["OdlServerToken"] = {
        "type": "apiKey",
        "in": "header",
        "name": "X-ODL-Token",
        "description": "Server token from Render Environment.",
    }
    schema["security"] = [{"OdlServerToken": []}]
    return JSONResponse(schema)


@app.get("/docs", include_in_schema=False)
def swagger_docs(request: Request) -> Any:
    auth_err = require_auth_token(request)
    if auth_err:
        return auth_err
    token = request.query_params.get("token")
    openapi_url = f"/openapi.json?token={token}" if token else "/openapi.json"
    return get_swagger_ui_html(openapi_url=openapi_url, title=f"{app.title} - Swagger")
